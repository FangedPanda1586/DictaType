from __future__ import annotations

import csv
import html
import json
import os
import queue
import subprocess
import sys
import time
import tkinter as tk
import tkinter.font as tkfont
import webbrowser
from datetime import datetime
from pathlib import Path
from tkinter import filedialog, messagebox, simpledialog, ttk
from typing import Any, Callable

from .classroom import ClassroomServer
from .db import Database, app_data_dir
from .performance import apply_runtime_hints, resolve_performance_profile
from .reporting import save_attempt_pdf, save_exam_pdf
from .scoring import calculate_wpm, score_text, split_sentences
from .tts import BUNDLED_FRENCH_VOICE_ID, SpeechEngine, Voice, builtin_french_available, french_voice_diagnostics, list_voices, verbalize_punctuation

APP_TITLE = "DictaType"
APP_VERSION = "1.0.0"

DARK = {
    "bg": "#20262E",
    "panel": "#2B333D",
    "panel2": "#36414D",
    "field": "#252D36",
    "border": "#465360",
    "text": "#F5F8FB",
    "muted": "#B1BDC8",
    "accent": "#8CC8FF",
    "accent2": "#B7DEFF",
    "accent_strong": "#4A9DE8",
    "accent_soft": "#D9EEFF",
    "danger": "#E36D78",
    "success": "#61BFA3",
    "sidebar": "#2A3139",
}

LIGHT = {
    "bg": "#F2F4F7",
    "panel": "#FFFFFF",
    "panel2": "#E8EDF2",
    "field": "#F8FAFC",
    "border": "#D2DAE3",
    "text": "#283746",
    "muted": "#718096",
    "accent": "#A9D8FF",
    "accent2": "#D9EEFF",
    "accent_strong": "#4B9FE6",
    "accent_soft": "#E7F4FF",
    "danger": "#D95D69",
    "success": "#42A982",
    "sidebar": "#E3E7EB",
}



def format_duration(seconds: int) -> str:
    seconds = max(0, int(seconds))
    minutes, secs = divmod(seconds, 60)
    hours, minutes = divmod(minutes, 60)
    if hours:
        return f"{hours:d}:{minutes:02d}:{secs:02d}"
    return f"{minutes:02d}:{secs:02d}"


def word_count(text: str) -> int:
    return len(text.split())


def open_folder(path: Path) -> None:
    path = path.resolve()
    if sys.platform == "win32":
        os.startfile(path)  # type: ignore[attr-defined]
    elif sys.platform == "darwin":
        subprocess.Popen(["open", str(path)])
    else:
        subprocess.Popen(["xdg-open", str(path)])


class NoPasteText(tk.Text):
    def __init__(self, master=None, **kwargs):
        super().__init__(master, **kwargs)
        self.bind("<<Paste>>", lambda _event: "break")
        self.bind("<Control-v>", lambda _event: "break")
        self.bind("<Control-V>", lambda _event: "break")
        self.bind("<Shift-Insert>", lambda _event: "break")
        self.bind("<Button-3>", lambda _event: "break")


class RoundedButton(tk.Canvas):
    """Small native-Tk rounded button used throughout the revamped interface."""

    def __init__(
        self,
        master,
        text: str = "",
        command: Callable[[], None] | None = None,
        style: str = "Secondary.TButton",
        state: str = "normal",
        width: int | None = None,
        **kwargs,
    ):
        self._app = self._find_app(master)
        self._text = text
        self._command = command
        self._style_name = style
        self._state = state
        self._selected = False
        self._hover = False
        self._font = tkfont.Font(family="Segoe UI", size=10, weight="bold")
        requested_width = max(104, (width or 0) * 8) if width else max(104, self._font.measure(text) + 34)
        bg = self._parent_background(master)
        super().__init__(
            master,
            width=requested_width,
            height=40,
            highlightthickness=0,
            bd=0,
            relief="flat",
            bg=bg,
            cursor="hand2" if state != "disabled" else "arrow",
            takefocus=1,
            **kwargs,
        )
        self.bind("<Configure>", lambda _e: self._draw())
        self.bind("<Enter>", self._enter)
        self.bind("<Leave>", self._leave)
        self.bind("<Button-1>", self._click)
        self.bind("<Return>", self._click)
        self.bind("<space>", self._click)
        self._draw()

    @staticmethod
    def _find_app(widget):
        current = widget
        while current is not None:
            if hasattr(current, "colors"):
                return current
            current = getattr(current, "master", None)
        return None

    def _parent_background(self, master) -> str:
        colors = self._app.colors if self._app else LIGHT
        try:
            style = master.cget("style") if isinstance(master, ttk.Widget) else ""
        except Exception:
            style = ""
        if style == "Card.TFrame":
            return colors["panel"]
        if style == "Sidebar.TFrame":
            return colors["sidebar"]
        if style == "Page.TFrame":
            return colors["bg"]
        if style == "Field.TFrame":
            return colors["field"]
        if isinstance(master, (ttk.Frame, ttk.LabelFrame)):
            return colors["panel"]
        try:
            return master.cget("bg")
        except Exception:
            return colors["bg"]

    def _palette(self) -> tuple[str, str]:
        c = self._app.colors if self._app else LIGHT
        if self._state == "disabled":
            return c["panel2"], c["muted"]
        if self._style_name == "Accent.TButton":
            return (c["accent"] if self._hover else c["accent_strong"]), "#FFFFFF"
        if self._style_name == "Danger.TButton":
            return ("#E97882" if self._hover else c["danger"]), "#FFFFFF"
        if self._style_name == "Sidebar.TButton":
            if self._selected:
                return c["accent_soft"], c["accent_strong"]
            return (c["panel2"] if self._hover else c["sidebar"]), c["text"]
        return ("#DDE5EC" if self._hover else c["panel2"]), c["text"]

    def _round_rect(self, x1, y1, x2, y2, radius, fill):
        radius = min(radius, max(2, (y2 - y1) / 2), max(2, (x2 - x1) / 2))
        self.create_rectangle(x1 + radius, y1, x2 - radius, y2, fill=fill, outline="")
        self.create_rectangle(x1, y1 + radius, x2, y2 - radius, fill=fill, outline="")
        self.create_oval(x1, y1, x1 + radius * 2, y1 + radius * 2, fill=fill, outline="")
        self.create_oval(x2 - radius * 2, y1, x2, y1 + radius * 2, fill=fill, outline="")
        self.create_oval(x1, y2 - radius * 2, x1 + radius * 2, y2, fill=fill, outline="")
        self.create_oval(x2 - radius * 2, y2 - radius * 2, x2, y2, fill=fill, outline="")

    def _draw(self):
        self.delete("all")
        width = max(2, self.winfo_width())
        height = max(2, self.winfo_height())
        fill, fg = self._palette()
        self._round_rect(1, 1, width - 1, height - 1, 11, fill)
        self.create_text(width / 2, height / 2, text=self._text, fill=fg, font=self._font)

    def _enter(self, _event=None):
        if self._state != "disabled":
            # Low-memory mode avoids unnecessary canvas redraws on older Intel
            # HD graphics while keeping the same rounded visual design.
            if getattr(getattr(self._app, "performance", None), "low_memory", False):
                return
            self._hover = True
            self._draw()

    def _leave(self, _event=None):
        if getattr(getattr(self._app, "performance", None), "low_memory", False):
            return
        self._hover = False
        self._draw()

    def _click(self, _event=None):
        if self._state != "disabled" and self._command:
            self._command()

    def configure(self, cnf=None, **kwargs):
        custom = False
        if "text" in kwargs:
            self._text = str(kwargs.pop("text"))
            custom = True
        if "state" in kwargs:
            self._state = str(kwargs.pop("state"))
            super().configure(cursor="arrow" if self._state == "disabled" else "hand2")
            custom = True
        if "command" in kwargs:
            self._command = kwargs.pop("command")
            custom = True
        if kwargs or cnf:
            result = super().configure(cnf, **kwargs)
        else:
            result = None
        if custom:
            self._draw()
        return result

    config = configure

    def state(self, statespec=None):
        if statespec is None:
            return ("selected",) if self._selected else ()
        for item in statespec:
            if item == "selected":
                self._selected = True
            elif item == "!selected":
                self._selected = False
            elif item == "disabled":
                self._state = "disabled"
            elif item == "!disabled":
                self._state = "normal"
        self._draw()
        return ()


class Card(ttk.Frame):
    def __init__(self, master, **kwargs):
        super().__init__(master, style="Card.TFrame", padding=kwargs.pop("padding", 18), **kwargs)


class PinDialog(simpledialog.Dialog):
    def __init__(self, parent, check: Callable[[str], bool]):
        self.check = check
        self.success = False
        super().__init__(parent, title="Teacher access")

    def body(self, master):
        ttk.Label(master, text="Enter the teacher PIN", style="DialogTitle.TLabel").grid(row=0, column=0, columnspan=2, sticky="w", pady=(0, 10))
        ttk.Label(master, text="PIN").grid(row=1, column=0, sticky="w", padx=(0, 10))
        self.entry = ttk.Entry(master, show="●", width=26)
        self.entry.grid(row=1, column=1, sticky="ew")
        master.columnconfigure(1, weight=1)
        return self.entry

    def validate(self):
        if self.check(self.entry.get()):
            self.success = True
            return True
        messagebox.showerror("Access denied", "The PIN is incorrect.", parent=self)
        self.entry.select_range(0, "end")
        return False


class LessonEditor(tk.Toplevel):
    def __init__(self, app: "DictaTypeApp", lesson: dict[str, Any] | None = None):
        super().__init__(app)
        self.app = app
        self.db = app.db
        self.lesson = lesson
        self.result_id: int | None = None
        self.title("Edit dictation" if lesson else "New dictation")
        self.geometry("900x800")
        self.minsize(780, 680)
        self.transient(app)
        self.grab_set()
        self.configure(bg=app.colors["bg"])
        self._build()
        self._load()
        self.protocol("WM_DELETE_WINDOW", self.destroy)

    def _build(self):
        outer = ttk.Frame(self, padding=20)
        outer.pack(fill="both", expand=True)
        outer.columnconfigure(1, weight=1)
        outer.rowconfigure(6, weight=1)

        ttk.Label(outer, text="Dictation details", style="PageTitle.TLabel").grid(row=0, column=0, columnspan=4, sticky="w", pady=(0, 16))

        ttk.Label(outer, text="Title").grid(row=1, column=0, sticky="w", pady=6)
        self.title_var = tk.StringVar()
        ttk.Entry(outer, textvariable=self.title_var).grid(row=1, column=1, columnspan=3, sticky="ew", pady=6)

        ttk.Label(outer, text="Language").grid(row=2, column=0, sticky="w", pady=6)
        self.language_var = tk.StringVar(value="English")
        self.language_combo = ttk.Combobox(outer, textvariable=self.language_var, values=["English", "Français"], state="readonly", width=18)
        self.language_combo.grid(row=2, column=1, sticky="ew", pady=6, padx=(0, 12))
        self.language_combo.bind("<<ComboboxSelected>>", lambda _e: self._refresh_voice_values())

        ttk.Label(outer, text="Difficulty").grid(row=2, column=2, sticky="w", pady=6)
        self.difficulty_var = tk.StringVar(value="Intermediate")
        ttk.Combobox(outer, textvariable=self.difficulty_var, values=["Beginner", "Intermediate", "Advanced"], state="readonly").grid(row=2, column=3, sticky="ew", pady=6)

        ttk.Label(outer, text="Category").grid(row=3, column=0, sticky="w", pady=6)
        self.category_var = tk.StringVar()
        ttk.Entry(outer, textvariable=self.category_var).grid(row=3, column=1, sticky="ew", pady=6, padx=(0, 12))

        ttk.Label(outer, text="Voice").grid(row=3, column=2, sticky="w", pady=6)
        self.voice_var = tk.StringVar()
        self.voice_combo = ttk.Combobox(outer, textvariable=self.voice_var, state="readonly")
        self.voice_combo.grid(row=3, column=3, sticky="ew", pady=6)

        ttk.Label(outer, text="Dictation type").grid(row=4, column=0, sticky="w", pady=(10, 6))
        self.mode_var = tk.StringVar(value="Sentence Mode")
        self.mode_combo = ttk.Combobox(
            outer,
            textvariable=self.mode_var,
            values=["Sentence Mode", "Passage Mode"],
            state="readonly",
            width=20,
        )
        self.mode_combo.grid(row=4, column=1, sticky="ew", pady=(10, 6), padx=(0, 12))
        self.mode_combo.bind("<<ComboboxSelected>>", lambda _e: self._mode_changed())
        self.mode_help = ttk.Label(outer, text="", style="Muted.TLabel", wraplength=430)
        self.mode_help.grid(row=4, column=2, columnspan=2, sticky="w", pady=(10, 6))

        ttk.Label(outer, text="Dictation text").grid(row=5, column=0, sticky="nw", pady=(12, 6))
        text_frame = ttk.Frame(outer, style="Field.TFrame")
        text_frame.grid(row=6, column=0, columnspan=4, sticky="nsew")
        text_frame.columnconfigure(0, weight=1)
        text_frame.rowconfigure(0, weight=1)
        self.text = tk.Text(
            text_frame,
            wrap="word",
            undo=True,
            font=("Segoe UI", 11),
            relief="flat",
            padx=12,
            pady=12,
            bg=self.app.colors["field"],
            fg=self.app.colors["text"],
            insertbackground=self.app.colors["text"],
            selectbackground=self.app.colors["accent"],
        )
        scroll = ttk.Scrollbar(text_frame, orient="vertical", command=self.text.yview)
        self.text.configure(yscrollcommand=scroll.set)
        self.text.grid(row=0, column=0, sticky="nsew")
        scroll.grid(row=0, column=1, sticky="ns")

        controls = ttk.LabelFrame(outer, text="Exercise controls", padding=14)
        controls.grid(row=7, column=0, columnspan=4, sticky="ew", pady=(14, 8))
        for i in range(4):
            controls.columnconfigure(i, weight=1)

        ttk.Label(controls, text="Speech rate").grid(row=0, column=0, sticky="w")
        self.rate_var = tk.IntVar(value=175)
        ttk.Spinbox(controls, from_=80, to=300, increment=5, textvariable=self.rate_var, width=10).grid(row=1, column=0, sticky="ew", padx=(0, 10))

        self.replay_label = ttk.Label(controls, text="Replay limit / sentence")
        self.replay_label.grid(row=0, column=1, sticky="w")
        self.replay_var = tk.IntVar(value=3)
        ttk.Spinbox(controls, from_=0, to=20, textvariable=self.replay_var, width=10).grid(row=1, column=1, sticky="ew", padx=(0, 10))

        ttk.Label(controls, text="Time limit (minutes, 0 = none)").grid(row=0, column=2, sticky="w")
        self.time_var = tk.IntVar(value=0)
        ttk.Spinbox(controls, from_=0, to=240, textvariable=self.time_var, width=10).grid(row=1, column=2, sticky="ew", padx=(0, 10))

        ttk.Label(controls, text="Marking mode").grid(row=0, column=3, sticky="w")
        self.marking_var = tk.StringVar(value="Balanced")
        ttk.Combobox(controls, textvariable=self.marking_var, values=["Flexible", "Balanced", "Strict"], state="readonly").grid(row=1, column=3, sticky="ew")

        self.punctuation_var = tk.BooleanVar(value=False)
        self.show_results_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(controls, text="Speak punctuation names", variable=self.punctuation_var).grid(row=2, column=2, sticky="w", pady=(12, 0))
        ttk.Checkbutton(controls, text="Show result to student", variable=self.show_results_var).grid(row=2, column=3, sticky="w", pady=(12, 0))

        buttons = ttk.Frame(outer)
        buttons.grid(row=8, column=0, columnspan=4, sticky="ew", pady=(10, 0))
        RoundedButton(buttons, text="Import text", style="Secondary.TButton", command=self.import_text).pack(side="left")
        RoundedButton(buttons, text="Preview voice", style="Secondary.TButton", command=self.preview).pack(side="left", padx=8)
        RoundedButton(buttons, text="Cancel", style="Secondary.TButton", command=self.destroy).pack(side="right")
        RoundedButton(buttons, text="Save dictation", style="Accent.TButton", command=self.save).pack(side="right", padx=8)

        self._mode_changed()

    def _mode_changed(self):
        sentence_mode = self.mode_var.get() == "Sentence Mode"
        if sentence_mode:
            self.mode_help.configure(text="Plays one sentence at a time and enables Previous / Next sentence controls for the student.")
            self.replay_label.configure(text="Replay limit / sentence")
        else:
            self.mode_help.configure(text="Plays the entire text continuously as one passage. Sentence navigation is hidden for the student.")
            self.replay_label.configure(text="Replay limit / passage")

    def _refresh_voice_values(self):
        language = "fr" if self.language_var.get() == "Français" else "en"
        matching = self.app.voices_for_language(language)
        values = [voice.display_name for voice in matching]
        self.voice_combo["values"] = values
        if values and self.voice_var.get() not in values:
            self.voice_var.set(values[0])
        if not values:
            self.voice_var.set("Use Windows default voice")
            self.voice_combo["values"] = ["Use Windows default voice"]

    def _load(self):
        self._refresh_voice_values()
        if not self.lesson:
            return
        self.title_var.set(self.lesson.get("title", ""))
        self.language_var.set("Français" if self.lesson.get("language") == "fr" else "English")
        self.difficulty_var.set(self.lesson.get("difficulty", "Intermediate"))
        self.category_var.set(self.lesson.get("category", ""))
        self.rate_var.set(int(self.lesson.get("rate", 175)))
        self.replay_var.set(int(self.lesson.get("replay_limit", 3)))
        self.time_var.set(int(self.lesson.get("time_limit", 0)) // 60)
        self.marking_var.set(str(self.lesson.get("marking_mode", "balanced")).title())
        self.mode_var.set("Sentence Mode" if bool(self.lesson.get("sentence_mode", 1)) else "Passage Mode")
        self._mode_changed()
        self.punctuation_var.set(bool(self.lesson.get("speak_punctuation", 0)))
        self.show_results_var.set(bool(self.lesson.get("show_results", 1)))
        self.text.insert("1.0", self.lesson.get("text", ""))
        self._refresh_voice_values()
        voice_name = self.lesson.get("voice_name", "")
        for voice in self.app.voices:
            if voice.name == voice_name or voice.id == self.lesson.get("voice_id", ""):
                self.voice_var.set(voice.display_name)
                break

    def selected_voice(self) -> Voice | None:
        display = self.voice_var.get()
        return next((voice for voice in self.app.voices if voice.display_name == display), None)

    def import_text(self):
        path = filedialog.askopenfilename(
            parent=self,
            title="Import passage",
            filetypes=[("Text or Word document", "*.txt *.docx"), ("Text file", "*.txt"), ("Word document", "*.docx")],
        )
        if not path:
            return
        try:
            source = Path(path)
            if source.suffix.lower() == ".txt":
                content = source.read_text(encoding="utf-8-sig")
            elif source.suffix.lower() == ".docx":
                from docx import Document

                document = Document(source)
                content = "\n".join(paragraph.text for paragraph in document.paragraphs if paragraph.text.strip())
            else:
                raise ValueError("Unsupported file format")
            self.text.delete("1.0", "end")
            self.text.insert("1.0", content)
            if not self.title_var.get().strip():
                self.title_var.set(source.stem)
        except Exception as exc:
            messagebox.showerror("Import failed", str(exc), parent=self)

    def preview(self):
        content = self.text.get("1.0", "end").strip()
        if not content:
            messagebox.showwarning("No passage", "Enter a passage before previewing it.", parent=self)
            return
        sample = split_sentences(content)[0] if split_sentences(content) else content[:300]
        language = "fr" if self.language_var.get() == "Français" else "en"
        if self.punctuation_var.get():
            sample = verbalize_punctuation(sample, language)
        voice = self.selected_voice()
        self.app.speech.speak(
            sample,
            voice_id=voice.id if voice else "",
            rate=self.rate_var.get(),
            on_error=lambda exc: self.app.post(lambda: messagebox.showerror("Speech error", str(exc), parent=self)),
        )

    def save(self):
        content = self.text.get("1.0", "end").strip()
        if not self.title_var.get().strip():
            messagebox.showwarning("Missing title", "Enter a title for the dictation.", parent=self)
            return
        if not content:
            messagebox.showwarning("Missing passage", "Enter the text to be dictated.", parent=self)
            return
        voice = self.selected_voice()
        data = {
            "title": self.title_var.get(),
            "language": "fr" if self.language_var.get() == "Français" else "en",
            "difficulty": self.difficulty_var.get(),
            "category": self.category_var.get(),
            "text": content,
            "voice_id": voice.id if voice else "",
            "voice_name": voice.name if voice else "",
            "rate": self.rate_var.get(),
            "volume": 1.0,
            "replay_limit": self.replay_var.get(),
            "time_limit": self.time_var.get() * 60,
            "marking_mode": self.marking_var.get().lower(),
            "sentence_mode": self.mode_var.get() == "Sentence Mode",
            "speak_punctuation": self.punctuation_var.get(),
            "show_results": self.show_results_var.get(),
        }
        try:
            self.result_id = self.db.save_lesson(data, self.lesson.get("id") if self.lesson else None)
            self.app.refresh_all()
            self.destroy()
        except Exception as exc:
            messagebox.showerror("Save failed", str(exc), parent=self)


class ResultDialog(tk.Toplevel):
    def __init__(self, app: "DictaTypeApp", attempt: dict[str, Any], read_only: bool = False):
        super().__init__(app)
        self.app = app
        self.attempt = attempt
        self.read_only = read_only
        self.title(f"Result · {attempt.get('student_name', '')}")
        self.geometry("920x720")
        self.minsize(780, 600)
        self.transient(app)
        self.configure(bg=app.colors["bg"])
        self._build()

    def _build(self):
        outer = ttk.Frame(self, padding=20)
        outer.pack(fill="both", expand=True)
        header = ttk.Frame(outer)
        header.pack(fill="x")
        ttk.Label(header, text=self.attempt.get("student_name", "Student"), style="PageTitle.TLabel").pack(side="left")
        ttk.Label(header, text=self.attempt.get("lesson_title", ""), style="Muted.TLabel").pack(side="right")

        # Keep export controls near the top so they stay visible on small screens.
        export_bar = ttk.Frame(outer)
        export_bar.pack(fill="x", pady=(10, 0))
        RoundedButton(
            export_bar,
            text="Save PDF",
            style="Accent.TButton",
            command=self.save_pdf_report,
        ).pack(side="left")
        RoundedButton(
            export_bar,
            text="Save HTML",
            style="Secondary.TButton",
            command=self.save_report,
        ).pack(side="left", padx=(8, 0))
        if not self.read_only and self.app.current_role == "teacher":
            RoundedButton(
                export_bar,
                text="Full student history",
                style="Secondary.TButton",
                command=self.view_history,
            ).pack(side="left", padx=(8, 0))
        exam_title = self.attempt.get("exam_title", "")
        if exam_title:
            item = ""
            if self.attempt.get("exam_item_count", 0):
                item = f" · Passage {self.attempt.get('exam_item_index', 0)} of {self.attempt.get('exam_item_count', 0)}"
            ttk.Label(outer, text=f"{exam_title}{item}", style="Muted.TLabel").pack(anchor="w", pady=(8, 0))

        metrics = ttk.Frame(outer)
        metrics.pack(fill="x", pady=16)
        values = [
            ("Overall", f"{self.attempt.get('overall_score', 0):.1f}%"),
            ("Words", f"{self.attempt.get('score_word', 0):.1f}%"),
            ("Characters", f"{self.attempt.get('score_char', 0):.1f}%"),
            ("Speed", f"{self.attempt.get('wpm', 0):.1f} WPM"),
            ("Time", format_duration(self.attempt.get("duration_seconds", 0))),
            ("Replays", str(self.attempt.get("replay_count", 0))),
        ]
        for index, (label, value) in enumerate(values):
            metrics.columnconfigure(index, weight=1)
            card = Card(metrics, padding=12)
            card.grid(row=0, column=index, sticky="ew", padx=(0 if index == 0 else 5, 0))
            ttk.Label(card, text=label, style="Muted.TLabel").pack(anchor="w")
            ttk.Label(card, text=value, style="Metric.TLabel").pack(anchor="w")

        notebook = ttk.Notebook(outer)
        notebook.pack(fill="both", expand=True)
        answer_tab = ttk.Frame(notebook, padding=12)
        changes_tab = ttk.Frame(notebook, padding=12)
        details_tab = ttk.Frame(notebook, padding=12)
        notebook.add(answer_tab, text="Student answer")
        notebook.add(changes_tab, text="Corrections")
        notebook.add(details_tab, text="Analysis")

        answer_text = tk.Text(answer_tab, wrap="word", relief="flat", padx=12, pady=12, bg=self.app.colors["field"], fg=self.app.colors["text"], insertbackground=self.app.colors["text"])
        answer_text.pack(fill="both", expand=True)
        answer_text.insert("1.0", self.attempt.get("answer", ""))
        answer_text.configure(state="disabled")

        changes_text = tk.Text(changes_tab, wrap="word", relief="flat", padx=12, pady=12, bg=self.app.colors["field"], fg=self.app.colors["text"], insertbackground=self.app.colors["text"])
        changes_text.pack(fill="both", expand=True)
        details = self.attempt.get("details", {})
        changes = details.get("changes", []) if isinstance(details, dict) else []
        if changes:
            for change in changes:
                kind = change.get("kind", "change").title()
                expected = change.get("expected", "") or "∅"
                actual = change.get("actual", "") or "∅"
                changes_text.insert("end", f"{kind}: {expected}  →  {actual}\n")
        else:
            changes_text.insert("end", "No word-level corrections. Excellent work.")
        changes_text.configure(state="disabled")

        analysis = [
            ("Marking mode", details.get("mode", "balanced")),
            ("Correct words", details.get("correct_words", 0)),
            ("Substitutions", details.get("substitutions", 0)),
            ("Missing words", details.get("missing_words", 0)),
            ("Extra words", details.get("extra_words", 0)),
            ("Accent mistakes", details.get("accent_mistakes", 0)),
            ("Capitalisation mistakes", details.get("capitalization_mistakes", 0)),
            ("Punctuation mistakes", details.get("punctuation_mistakes", 0)),
        ]
        for row, (label, value) in enumerate(analysis):
            ttk.Label(details_tab, text=label, style="Muted.TLabel").grid(row=row, column=0, sticky="w", padx=(0, 20), pady=6)
            ttk.Label(details_tab, text=str(value)).grid(row=row, column=1, sticky="w", pady=6)

        comment_frame = ttk.LabelFrame(outer, text="Teacher comment", padding=10)
        comment_frame.pack(fill="x", pady=(14, 0))
        self.comment_var = tk.StringVar(value=self.attempt.get("teacher_comment", ""))
        if self.read_only:
            comment = self.comment_var.get().strip() or "No teacher comment has been added."
            ttk.Label(comment_frame, text=comment, wraplength=760).pack(anchor="w", fill="x")
        else:
            ttk.Entry(comment_frame, textvariable=self.comment_var).pack(side="left", fill="x", expand=True)
            RoundedButton(comment_frame, text="Save comment", style="Secondary.TButton", command=self.save_comment).pack(side="left", padx=(10, 0))

    def view_history(self):
        student = self.app.db.get_student(self.attempt.get("student_id")) if self.attempt.get("student_id") else None
        if student is None:
            student = {"id": None, "name": self.attempt.get("student_name", "Student"), "class_name": self.attempt.get("class_name", ""), "identifier": ""}
        StudentHistoryDialog(self.app, student)

    def save_comment(self):
        self.app.db.update_attempt_comment(int(self.attempt["id"]), self.comment_var.get())
        self.attempt["teacher_comment"] = self.comment_var.get()
        self.app.refresh_all()
        messagebox.showinfo("Saved", "The teacher comment was saved.", parent=self)

    def save_pdf_report(self):
        student_name = str(self.attempt.get("student_name", "result")).strip() or "result"
        safe_name = "".join(
            char if char.isalnum() or char in {"-", "_", " "} else "_"
            for char in student_name
        ).strip().replace(" ", "_") or "result"
        path = filedialog.asksaveasfilename(
            parent=self,
            defaultextension=".pdf",
            filetypes=[("PDF document", "*.pdf")],
            initialfile=f"DictaType_{safe_name}_analysis.pdf",
        )
        if not path:
            return
        try:
            save_attempt_pdf(
                self.attempt,
                path,
                teacher_comment=self.comment_var.get(),
            )
        except Exception as exc:
            messagebox.showerror(
                "PDF export failed",
                f"The PDF could not be created.\n\n{exc}",
                parent=self,
            )
            return
        messagebox.showinfo(
            "PDF saved",
            "The assessment analysis was saved as a PDF.",
            parent=self,
        )

    def save_report(self):
        path = filedialog.asksaveasfilename(parent=self, defaultextension=".html", filetypes=[("HTML report", "*.html")], initialfile=f"DictaType_{self.attempt.get('student_name','result')}.html")
        if not path:
            return
        details = self.attempt.get("details", {})
        changes = details.get("changes", []) if isinstance(details, dict) else []
        rows = "".join(
            f"<tr><td>{html.escape(str(item.get('kind','')))}</td><td>{html.escape(str(item.get('expected','')))}</td><td>{html.escape(str(item.get('actual','')))}</td></tr>"
            for item in changes
        ) or "<tr><td colspan='3'>No word-level corrections.</td></tr>"
        document = f"""<!doctype html><meta charset='utf-8'><title>DictaType Result</title>
<style>body{{font:16px system-ui;max-width:900px;margin:40px auto;color:#1e2943}}.metrics{{display:flex;gap:12px;flex-wrap:wrap}}.box{{background:#eef2fb;padding:15px 20px;border-radius:12px}}table{{border-collapse:collapse;width:100%}}td,th{{border:1px solid #ccd4e5;padding:8px;text-align:left}}pre{{white-space:pre-wrap;background:#f5f7fc;padding:16px;border-radius:12px}}</style>
<h1>DictaType Result</h1><h2>{html.escape(self.attempt.get('student_name',''))}</h2><p>{html.escape(self.attempt.get('lesson_title',''))}</p>
<div class='metrics'><div class='box'><b>Overall</b><br>{self.attempt.get('overall_score',0):.1f}%</div><div class='box'><b>Word accuracy</b><br>{self.attempt.get('score_word',0):.1f}%</div><div class='box'><b>WPM</b><br>{self.attempt.get('wpm',0):.1f}</div><div class='box'><b>Time</b><br>{format_duration(self.attempt.get('duration_seconds',0))}</div></div>
<h3>Student answer</h3><pre>{html.escape(self.attempt.get('answer',''))}</pre><h3>Corrections</h3><table><tr><th>Type</th><th>Expected</th><th>Typed</th></tr>{rows}</table><h3>Teacher comment</h3><p>{html.escape(self.comment_var.get())}</p>"""
        Path(path).write_text(document, encoding="utf-8")
        messagebox.showinfo("Report saved", "The HTML report was created.", parent=self)


class StudentHistoryDialog(tk.Toplevel):
    """Teacher-only view of every stored result for one student profile."""

    def __init__(self, app: "DictaTypeApp", student: dict[str, Any]):
        super().__init__(app)
        self.app = app
        self.db = app.db
        self.student = student
        self.title(f"Result history · {student.get('name', 'Student')}")
        self.geometry("1040x680")
        self.minsize(820, 560)
        self.transient(app)
        self.configure(bg=app.colors["bg"])
        self._build()
        self.refresh()

    def _build(self):
        outer = ttk.Frame(self, padding=20)
        outer.pack(fill="both", expand=True)
        header = ttk.Frame(outer)
        header.pack(fill="x")
        ttk.Label(header, text=self.student.get("name", "Student"), style="PageTitle.TLabel").pack(side="left")
        profile = self.student.get("class_name", "") or "No class"
        identifier = self.student.get("identifier", "")
        if identifier:
            profile += f" · {identifier}"
        ttk.Label(header, text=profile, style="RoleBadge.TLabel").pack(side="right")
        ttk.Label(outer, text="All previous DictaType attempts are kept here, including classroom exam passages.", style="Muted.TLabel").pack(anchor="w", pady=(4, 14))

        metrics = ttk.Frame(outer)
        metrics.pack(fill="x", pady=(0, 14))
        self.metric_labels = {}
        for index, (key, title) in enumerate([("count", "Attempts"), ("average", "Average"), ("best", "Best score"), ("latest", "Latest")]):
            metrics.columnconfigure(index, weight=1)
            card = Card(metrics, padding=12)
            card.grid(row=0, column=index, sticky="ew", padx=(0 if index == 0 else 5, 5 if index < 3 else 0))
            ttk.Label(card, text=title, style="CardMuted.TLabel").pack(anchor="w")
            value = ttk.Label(card, text="—", style="Metric.TLabel")
            value.pack(anchor="w", pady=(3, 0))
            self.metric_labels[key] = value

        bar = ttk.Frame(outer)
        bar.pack(fill="x", pady=(0, 10))
        RoundedButton(bar, text="View selected result", style="Accent.TButton", command=self.view_selected).pack(side="left")
        RoundedButton(bar, text="Export history to Excel", style="Secondary.TButton", command=self.export_xlsx).pack(side="left", padx=8)

        frame = ttk.Frame(outer)
        frame.pack(fill="both", expand=True)
        frame.columnconfigure(0, weight=1)
        frame.rowconfigure(0, weight=1)
        columns = ("date", "exam", "lesson", "score", "wpm", "source")
        self.tree = ttk.Treeview(frame, columns=columns, show="headings", selectmode="browse")
        specs = [
            ("date", "Date", 150),
            ("exam", "Exam / session", 210),
            ("lesson", "Dictation / passage", 270),
            ("score", "Score", 85),
            ("wpm", "WPM", 70),
            ("source", "Source", 110),
        ]
        for column, title, width in specs:
            self.tree.heading(column, text=title)
            self.tree.column(column, width=width, anchor="w")
        self.tree.grid(row=0, column=0, sticky="nsew")
        self.tree.bind("<Double-1>", lambda _e: self.view_selected())
        scroll = ttk.Scrollbar(frame, command=self.tree.yview)
        scroll.grid(row=0, column=1, sticky="ns")
        self.tree.configure(yscrollcommand=scroll.set)

    def _attempts(self) -> list[dict[str, Any]]:
        return self.db.list_student_attempts(self.student.get("id"), self.student.get("name", ""))

    def refresh(self):
        attempts = self._attempts()
        summary = self.db.student_history_summary(self.student.get("id"), self.student.get("name", ""))
        self.metric_labels["count"].configure(text=str(summary["attempt_count"]))
        self.metric_labels["average"].configure(text=f"{summary['average_score']:.1f}%")
        self.metric_labels["best"].configure(text=f"{summary['best_score']:.1f}%")
        latest = str(summary.get("latest_at", "")).replace("T", " ")[:10] or "—"
        self.metric_labels["latest"].configure(text=latest)
        self.tree.delete(*self.tree.get_children())
        for item in attempts:
            exam = item.get("exam_title", "") or "Practice"
            if item.get("exam_item_count", 0):
                exam += f" ({item.get('exam_item_index', 0)}/{item.get('exam_item_count', 0)})"
            self.tree.insert(
                "", "end", iid=str(item["id"]),
                values=(
                    item.get("created_at", "").replace("T", " ")[:16],
                    exam,
                    item.get("lesson_title", ""),
                    f"{item.get('overall_score', 0):.1f}%",
                    f"{item.get('wpm', 0):.1f}",
                    item.get("source", ""),
                ),
            )

    def view_selected(self):
        selection = self.tree.selection()
        if not selection:
            return
        attempt = self.db.get_attempt(int(selection[0]))
        if attempt:
            ResultDialog(self.app, attempt)

    def export_xlsx(self):
        attempts = self._attempts()
        if not attempts:
            messagebox.showinfo("No results", "This student does not have any saved attempts yet.", parent=self)
            return
        safe = "".join(ch if ch.isalnum() or ch in "-_ " else "_" for ch in self.student.get("name", "Student")).strip().replace(" ", "_")
        path = filedialog.asksaveasfilename(parent=self, defaultextension=".xlsx", filetypes=[("Excel workbook", "*.xlsx")], initialfile=f"DictaType_{safe}_history.xlsx")
        if not path:
            return
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            from openpyxl.utils import get_column_letter
            wb = Workbook()
            ws = wb.active
            ws.title = "Student history"
            headers = ["Date", "Exam / Session", "Item", "Dictation / Passage", "Overall Score (%)", "Word Accuracy (%)", "Character Accuracy (%)", "WPM", "Duration (seconds)", "Replays", "Source", "Teacher Comment"]
            ws.append(headers)
            for item in attempts:
                ws.append([
                    item.get("created_at"), item.get("exam_title", ""),
                    f"{item.get('exam_item_index', 0)}/{item.get('exam_item_count', 0)}" if item.get("exam_item_count", 0) else "",
                    item.get("lesson_title"), item.get("overall_score"), item.get("score_word"), item.get("score_char"), item.get("wpm"), item.get("duration_seconds"), item.get("replay_count"), item.get("source"), item.get("teacher_comment"),
                ])
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="4B9FE6")
            widths = [22, 26, 10, 34, 18, 18, 20, 10, 18, 10, 14, 34]
            for index, width in enumerate(widths, start=1):
                ws.column_dimensions[get_column_letter(index)].width = width
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            wb.save(path)
            messagebox.showinfo("Export complete", "The complete student history was exported.", parent=self)
        except Exception as exc:
            messagebox.showerror("Export failed", str(exc), parent=self)


class StudentPage(ttk.Frame):
    def __init__(self, app: "DictaTypeApp", master):
        super().__init__(master, padding=24, style="Page.TFrame")
        self.app = app
        self.db = app.db
        self.current_lesson: dict[str, Any] | None = None
        self.sentences: list[str] = []
        self.sentence_index = 0
        self.play_counts: dict[int, int] = {}
        self.started_at = 0.0
        self.timer_job = None
        self.lesson_map: dict[str, int] = {}
        self.student_var = tk.StringVar()
        self.class_var = tk.StringVar()
        self._build()

    def _build(self):
        self.columnconfigure(0, weight=1)
        self.rowconfigure(3, weight=1)

        header = ttk.Frame(self)
        header.grid(row=0, column=0, sticky="ew")
        header.columnconfigure(0, weight=1)
        ttk.Label(header, text="Your dictation workspace", style="PageTitle.TLabel").grid(row=0, column=0, sticky="w")
        self.profile_badge = ttk.Label(header, text="", style="RoleBadge.TLabel")
        self.profile_badge.grid(row=0, column=1, sticky="e")
        ttk.Label(
            self,
            text="Choose a dictation, listen carefully and type exactly what you hear.",
            style="Muted.TLabel",
        ).grid(row=1, column=0, sticky="w", pady=(4, 16))

        selector = Card(self, padding=16)
        selector.grid(row=2, column=0, sticky="ew", pady=(0, 14))
        selector.columnconfigure(1, weight=1)
        ttk.Label(selector, text="Dictation", style="SectionTitle.TLabel").grid(row=0, column=0, sticky="w", padx=(0, 16))
        self.lesson_var = tk.StringVar()
        self.lesson_combo = ttk.Combobox(selector, textvariable=self.lesson_var, state="readonly")
        self.lesson_combo.grid(row=0, column=1, sticky="ew", padx=(0, 12))
        self.start_button = RoundedButton(selector, text="Start exercise", style="Accent.TButton", command=self.start_exercise)
        self.start_button.grid(row=0, column=2, sticky="e")

        self.setup_card = Card(self, padding=18)
        self.setup_card.grid(row=3, column=0, sticky="nsew")
        self.setup_card.columnconfigure(0, weight=1)
        self.setup_card.rowconfigure(0, weight=1)

        exercise = ttk.Frame(self.setup_card, style="Card.TFrame")
        exercise.grid(row=0, column=0, sticky="nsew")
        exercise.columnconfigure(0, weight=1)
        exercise.rowconfigure(3, weight=1)

        top = ttk.Frame(exercise, style="Card.TFrame")
        top.grid(row=0, column=0, sticky="ew")
        self.exercise_title = ttk.Label(top, text="Ready when you are", style="CardSectionTitle.TLabel")
        self.exercise_title.pack(side="left")
        self.timer_label = ttk.Label(top, text="00:00", style="CardTimer.TLabel")
        self.timer_label.pack(side="right")

        self.progress_label = ttk.Label(exercise, text="Select a dictation above to begin.", style="CardMuted.TLabel")
        self.progress_label.grid(row=1, column=0, sticky="w", pady=(5, 12))

        action_bar = ttk.Frame(exercise, style="Card.TFrame")
        action_bar.grid(row=2, column=0, sticky="ew", pady=(0, 12))
        self.listen_button = RoundedButton(action_bar, text="▶ Listen", style="Accent.TButton", command=self.play_current, state="disabled")
        self.listen_button.pack(side="left")
        self.previous_button = RoundedButton(action_bar, text="Previous", style="Secondary.TButton", command=self.previous_sentence, state="disabled")
        self.previous_button.pack(side="left", padx=(8, 0))
        self.next_button = RoundedButton(action_bar, text="Next sentence", style="Secondary.TButton", command=self.next_sentence, state="disabled")
        self.next_button.pack(side="left", padx=(8, 0))
        self.fullscreen_button = RoundedButton(action_bar, text="Full screen", style="Secondary.TButton", command=self.toggle_fullscreen, state="disabled")
        self.fullscreen_button.pack(side="right")

        answer_frame = ttk.Frame(exercise, style="Field.TFrame")
        answer_frame.grid(row=3, column=0, sticky="nsew")
        answer_frame.columnconfigure(0, weight=1)
        answer_frame.rowconfigure(0, weight=1)
        self.answer = NoPasteText(
            answer_frame,
            wrap="word",
            undo=False,
            relief="flat",
            padx=18,
            pady=18,
            font=("Segoe UI", 12),
            bg=self.app.colors["field"],
            fg=self.app.colors["text"],
            insertbackground=self.app.colors["text"],
            selectbackground=self.app.colors["accent"],
            state="disabled",
        )
        self.answer.grid(row=0, column=0, sticky="nsew")
        scroll = ttk.Scrollbar(answer_frame, command=self.answer.yview)
        scroll.grid(row=0, column=1, sticky="ns")
        self.answer.configure(yscrollcommand=scroll.set)

        bottom = ttk.Frame(exercise, style="Card.TFrame")
        bottom.grid(row=4, column=0, sticky="ew", pady=(12, 0))
        self.status_label = ttk.Label(bottom, text="The original text stays hidden during the exercise.", style="CardMuted.TLabel")
        self.status_label.pack(side="left")
        self.cancel_button = RoundedButton(bottom, text="Cancel", style="Secondary.TButton", command=self.cancel_exercise, state="disabled")
        self.cancel_button.pack(side="right")
        self.submit_button = RoundedButton(bottom, text="Submit answer", style="Accent.TButton", command=self.submit, state="disabled")
        self.submit_button.pack(side="right", padx=(0, 8))

    def refresh(self):
        student = self.app.current_student or {}
        self.student_var.set(student.get("name", ""))
        self.class_var.set(student.get("class_name", ""))
        profile = student.get("name", "Student")
        if student.get("class_name"):
            profile += f"  ·  {student['class_name']}"
        self.profile_badge.configure(text=profile)

        lessons = self.db.list_lessons()
        self.lesson_map = {}
        lesson_values = []
        for lesson in lessons:
            lang = "FR" if lesson.get("language") == "fr" else "EN"
            mode = "Sentence" if lesson.get("sentence_mode", 1) else "Passage"
            label = f"{lesson['title']}  ·  {lang}  ·  {mode}  ·  {lesson['difficulty']}"
            if label in self.lesson_map:
                label += f"  #{lesson['id']}"
            self.lesson_map[label] = lesson["id"]
            lesson_values.append(label)
        self.lesson_combo["values"] = lesson_values
        if lesson_values and self.lesson_var.get() not in lesson_values:
            self.lesson_var.set(lesson_values[0])

    def start_exercise(self):
        student = self.app.current_student
        lesson_id = self.lesson_map.get(self.lesson_var.get())
        if not student:
            messagebox.showerror("Session expired", "Please sign in again before starting an exercise.", parent=self)
            self.app.logout()
            return
        if not lesson_id:
            messagebox.showwarning("Dictation", "Select a dictation.", parent=self)
            return
        lesson = self.db.get_lesson(lesson_id)
        if not lesson:
            messagebox.showerror("Missing dictation", "The selected dictation could not be found.", parent=self)
            return
        self.current_lesson = lesson
        is_sentence_mode = bool(lesson.get("sentence_mode", 1))
        self.sentences = split_sentences(lesson["text"]) if is_sentence_mode else [lesson["text"]]
        self.sentence_index = 0
        self.play_counts = {}
        self.started_at = time.monotonic()
        self.answer.configure(state="normal")
        self.answer.delete("1.0", "end")
        self.lesson_combo.configure(state="disabled")
        self.start_button.configure(state="disabled")
        self.listen_button.configure(state="normal", text="▶ Play sentence" if is_sentence_mode else "▶ Play passage")
        self.cancel_button.configure(state="normal")
        self.submit_button.configure(state="normal", text="Submit answer" if is_sentence_mode else "Submit passage")
        self.fullscreen_button.configure(state="normal")
        self.exercise_title.configure(text=lesson["title"])
        if is_sentence_mode:
            if not self.previous_button.winfo_manager():
                self.previous_button.pack(side="left", padx=(8, 0), after=self.listen_button)
            if not self.next_button.winfo_manager():
                self.next_button.pack(side="left", padx=(8, 0), after=self.previous_button)
            self.status_label.configure(text="Sentence mode: listen one sentence at a time. Copy and paste are disabled.")
        else:
            self.previous_button.pack_forget()
            self.next_button.pack_forget()
            self.status_label.configure(text="Passage mode: the entire passage plays continuously. Copy and paste are disabled.")
        self._update_progress()
        self._tick()
        self.after(350, self.play_current)
        self.answer.focus_set()

    def current_voice(self) -> Voice | None:
        if not self.current_lesson:
            return None
        voice_id = self.current_lesson.get("voice_id", "")
        return next((voice for voice in self.app.voices if voice.id == voice_id), None)

    def play_current(self):
        if not self.current_lesson or not self.sentences:
            return
        limit = int(self.current_lesson.get("replay_limit", 3))
        count = self.play_counts.get(self.sentence_index, 0)
        # One initial playback plus the configured number of replays.
        if limit > 0 and count >= limit + 1:
            unit = "sentence" if self.current_lesson.get("sentence_mode", 1) else "passage"
            self.status_label.configure(text=f"The replay limit for this {unit} has been reached.")
            return
        text = self.sentences[self.sentence_index]
        if self.current_lesson.get("speak_punctuation"):
            text = verbalize_punctuation(text, self.current_lesson.get("language", "en"))
        self.play_counts[self.sentence_index] = count + 1
        self.status_label.configure(text="Playing audio…")
        voice = self.current_voice()
        self.app.speech.speak(
            text,
            voice_id=voice.id if voice else self.current_lesson.get("voice_id", ""),
            rate=int(self.current_lesson.get("rate", 175)),
            volume=float(self.current_lesson.get("volume", 1.0)),
            on_done=lambda: self.app.post(lambda: self.status_label.configure(text="Audio finished. Continue typing.")),
            on_error=lambda exc: self.app.post(lambda: messagebox.showerror("Speech error", str(exc), parent=self)),
        )
        self._update_progress()

    def previous_sentence(self):
        if self.sentence_index > 0:
            self.sentence_index -= 1
            self._update_progress()

    def next_sentence(self):
        if self.sentence_index < len(self.sentences) - 1:
            self.sentence_index += 1
            self._update_progress()
            self.play_current()
        else:
            self.status_label.configure(text="This is the final sentence. Review your answer and submit.")

    def _update_progress(self):
        total = max(1, len(self.sentences))
        current = self.sentence_index + 1
        count = self.play_counts.get(self.sentence_index, 0)
        replay_limit = int(self.current_lesson.get("replay_limit", 3)) if self.current_lesson else 0
        replay_text = "Unlimited replays" if replay_limit == 0 else f"{max(0, count - 1)}/{replay_limit} replays"
        sentence_mode = bool(self.current_lesson.get("sentence_mode", 1)) if self.current_lesson else True
        if sentence_mode:
            self.progress_label.configure(text=f"Sentence {current} of {total}  ·  {replay_text}")
            self.previous_button.configure(state="normal" if self.sentence_index > 0 else "disabled")
            self.next_button.configure(state="normal" if self.sentence_index < total - 1 else "disabled")
        else:
            self.progress_label.configure(text=f"Passage mode  ·  {replay_text}")
            self.previous_button.configure(state="disabled")
            self.next_button.configure(state="disabled")

    def _tick(self):
        if not self.current_lesson:
            return
        elapsed = int(time.monotonic() - self.started_at)
        limit = int(self.current_lesson.get("time_limit", 0))
        if limit > 0:
            remaining = max(0, limit - elapsed)
            self.timer_label.configure(text=format_duration(remaining))
            if remaining <= 0:
                self.status_label.configure(text="Time is up. Your answer has been submitted automatically.")
                self.submit(auto=True)
                return
        else:
            self.timer_label.configure(text=format_duration(elapsed))
        self.timer_job = self.after(1000, self._tick)

    def total_plays(self) -> int:
        return sum(self.play_counts.values())

    def submit(self, auto: bool = False):
        if not self.current_lesson:
            return
        answer = self.answer.get("1.0", "end").strip()
        if not answer and not auto:
            if not messagebox.askyesno("Empty answer", "Submit an empty answer?", parent=self):
                return
        elapsed = max(1, int(time.monotonic() - self.started_at))
        result = score_text(self.current_lesson["text"], answer, self.current_lesson.get("marking_mode", "balanced"))
        wpm = calculate_wpm(answer, elapsed)
        student = self.app.current_student or {}
        student_name = student.get("name", "Student")
        attempt_id = self.db.save_attempt(
            {
                "student_id": student.get("id"),
                "student_name": student_name,
                "class_name": student.get("class_name", ""),
                "lesson_id": self.current_lesson["id"],
                "lesson_title": self.current_lesson["title"],
                "expected_text": self.current_lesson["text"],
                "answer": answer,
                "score_word": result.word_accuracy,
                "score_char": result.character_accuracy,
                "overall_score": result.overall_score,
                "wpm": wpm,
                "duration_seconds": elapsed,
                "replay_count": self.total_plays(),
                "details": result.to_dict(),
                "source": "desktop",
            }
        )
        show = bool(self.current_lesson.get("show_results", 1))
        self._finish_exercise()
        self.app.refresh_all()
        if show:
            attempt = self.db.get_attempt(attempt_id)
            if attempt:
                ResultDialog(self.app, attempt, read_only=True)
        else:
            messagebox.showinfo("Answer submitted", "The answer was saved for the teacher.", parent=self)

    def cancel_exercise(self):
        if self.current_lesson and not messagebox.askyesno("Cancel exercise", "Discard this attempt and return to the start screen?", parent=self):
            return
        self._finish_exercise()

    def _finish_exercise(self):
        self.app.speech.stop()
        if self.timer_job:
            try:
                self.after_cancel(self.timer_job)
            except Exception:
                pass
        self.timer_job = None
        self.current_lesson = None
        self.sentences = []
        self.answer.configure(state="normal")
        self.answer.delete("1.0", "end")
        self.answer.configure(state="disabled")
        self.lesson_combo.configure(state="readonly")
        self.start_button.configure(state="normal")
        for widget in [self.listen_button, self.previous_button, self.next_button, self.cancel_button, self.submit_button, self.fullscreen_button]:
            widget.configure(state="disabled")
        self.exercise_title.configure(text="Choose a dictation to begin")
        self.listen_button.configure(text="▶ Listen")
        self.submit_button.configure(text="Submit answer")
        self.progress_label.configure(text="")
        self.timer_label.configure(text="00:00")
        self.status_label.configure(text="The original passage stays hidden.")
        self.app.attributes("-fullscreen", False)

    def toggle_fullscreen(self):
        self.app.attributes("-fullscreen", not bool(self.app.attributes("-fullscreen")))


class TeacherPage(ttk.Frame):
    def __init__(self, app: "DictaTypeApp", master):
        super().__init__(master, padding=24, style="Page.TFrame")
        self.app = app
        self.db = app.db
        self._build()

    def _build(self):
        header = ttk.Frame(self)
        header.pack(fill="x")
        ttk.Label(header, text="Teacher workspace", style="PageTitle.TLabel").pack(side="left")
        ttk.Label(header, text="SECURE TEACHER SESSION", style="RoleBadge.TLabel").pack(side="right")
        ttk.Label(self, text="Manage dictations, student access and assessment results.", style="Muted.TLabel").pack(anchor="w", pady=(4, 14))

        metrics = ttk.Frame(self)
        metrics.pack(fill="x", pady=(0, 14))
        self.metric_labels: dict[str, ttk.Label] = {}
        for index, (key, label) in enumerate([
            ("dictations", "Dictations"),
            ("students", "Students"),
            ("attempts", "Attempts"),
            ("average", "Average score"),
        ]):
            metrics.columnconfigure(index, weight=1)
            card = Card(metrics, padding=14)
            card.grid(row=0, column=index, sticky="ew", padx=(0 if index == 0 else 5, 5 if index < 3 else 0))
            ttk.Label(card, text=label, style="CardMuted.TLabel").pack(anchor="w")
            value = ttk.Label(card, text="0", style="Metric.TLabel")
            value.pack(anchor="w", pady=(3, 0))
            self.metric_labels[key] = value

        self.notebook = ttk.Notebook(self)
        self.notebook.pack(fill="both", expand=True)
        self._build_lessons_tab()
        self._build_students_tab()
        self._build_results_tab()

    def _build_lessons_tab(self):
        tab = ttk.Frame(self.notebook, padding=14)
        self.notebook.add(tab, text="Dictations")
        tab.columnconfigure(0, weight=1)
        tab.rowconfigure(1, weight=1)
        bar = ttk.Frame(tab)
        bar.grid(row=0, column=0, sticky="ew", pady=(0, 10))
        RoundedButton(bar, text="New dictation", style="Accent.TButton", command=lambda: LessonEditor(self.app)).pack(side="left")
        RoundedButton(bar, text="Edit", style="Secondary.TButton", command=self.edit_lesson).pack(side="left", padx=6)
        RoundedButton(bar, text="Duplicate", style="Secondary.TButton", command=self.duplicate_lesson).pack(side="left")
        RoundedButton(bar, text="Delete", style="Danger.TButton", command=self.delete_lesson).pack(side="left", padx=6)
        RoundedButton(bar, text="Import JSON", style="Secondary.TButton", command=self.import_lessons_json).pack(side="right")
        RoundedButton(bar, text="Export JSON", style="Secondary.TButton", command=self.export_lessons_json).pack(side="right", padx=6)
        columns = ("title", "lang", "type", "category", "difficulty", "words", "updated")
        self.lessons_tree = ttk.Treeview(tab, columns=columns, show="headings", selectmode="browse")
        headings = {"title": "Title", "lang": "Language", "type": "Type", "category": "Category", "difficulty": "Difficulty", "words": "Words", "updated": "Updated"}
        widths = {"title": 250, "lang": 80, "type": 90, "category": 120, "difficulty": 105, "words": 65, "updated": 145}
        for column in columns:
            self.lessons_tree.heading(column, text=headings[column])
            self.lessons_tree.column(column, width=widths[column], anchor="w")
        self.lessons_tree.grid(row=1, column=0, sticky="nsew")
        self.lessons_tree.bind("<Double-1>", lambda _e: self.edit_lesson())
        scroll = ttk.Scrollbar(tab, command=self.lessons_tree.yview)
        scroll.grid(row=1, column=1, sticky="ns")
        self.lessons_tree.configure(yscrollcommand=scroll.set)

    def _build_students_tab(self):
        tab = ttk.Frame(self.notebook, padding=14)
        self.notebook.add(tab, text="Students")
        tab.columnconfigure(0, weight=1)
        tab.rowconfigure(1, weight=1)
        bar = ttk.Frame(tab)
        bar.grid(row=0, column=0, sticky="ew", pady=(0, 10))
        RoundedButton(bar, text="Add student", style="Accent.TButton", command=self.add_student).pack(side="left")
        RoundedButton(bar, text="Edit", style="Secondary.TButton", command=self.edit_student).pack(side="left", padx=6)
        RoundedButton(bar, text="View full history", style="Secondary.TButton", command=self.view_student_history).pack(side="left")
        RoundedButton(bar, text="Delete", style="Danger.TButton", command=self.delete_student).pack(side="left", padx=6)
        columns = ("name", "class", "identifier", "active", "created")
        self.students_tree = ttk.Treeview(tab, columns=columns, show="headings", selectmode="browse")
        for column, label, width in [("name", "Name", 240), ("class", "Class", 150), ("identifier", "Profile ID", 150), ("active", "Access", 100), ("created", "Added", 155)]:
            self.students_tree.heading(column, text=label)
            self.students_tree.column(column, width=width, anchor="w")
        self.students_tree.grid(row=1, column=0, sticky="nsew")
        self.students_tree.bind("<Double-1>", lambda _e: self.view_student_history())
        scroll = ttk.Scrollbar(tab, command=self.students_tree.yview)
        scroll.grid(row=1, column=1, sticky="ns")
        self.students_tree.configure(yscrollcommand=scroll.set)

    def _build_results_tab(self):
        tab = ttk.Frame(self.notebook, padding=14)
        self.notebook.add(tab, text="Results")
        tab.columnconfigure(0, weight=1)
        tab.rowconfigure(2, weight=1)
        filters = ttk.Frame(tab)
        filters.grid(row=0, column=0, sticky="ew", pady=(0, 8))
        ttk.Label(filters, text="Search").pack(side="left")
        self.result_search_var = tk.StringVar()
        entry = ttk.Entry(filters, textvariable=self.result_search_var, width=35)
        entry.pack(side="left", padx=(8, 14))
        entry.bind("<KeyRelease>", lambda _e: self.refresh_results())
        ttk.Label(filters, text="Search by student, class, passage or exam.", style="Muted.TLabel").pack(side="left")
        self.result_count_label = ttk.Label(filters, text="", style="Muted.TLabel")
        self.result_count_label.pack(side="right")
        bar = ttk.Frame(tab)
        bar.grid(row=1, column=0, sticky="ew", pady=(0, 10))
        RoundedButton(bar, text="View result", style="Accent.TButton", command=self.view_result).pack(side="left")
        RoundedButton(bar, text="Student history", style="Secondary.TButton", command=self.view_result_student_history).pack(side="left", padx=6)
        RoundedButton(bar, text="Full exam PDF", style="Secondary.TButton", command=self.save_selected_exam_report).pack(side="left")
        RoundedButton(bar, text="Delete", style="Danger.TButton", command=self.delete_result).pack(side="left", padx=6)
        RoundedButton(bar, text="Export Excel", style="Secondary.TButton", command=self.export_results_xlsx).pack(side="right")
        RoundedButton(bar, text="Export CSV", style="Secondary.TButton", command=self.export_results_csv).pack(side="right", padx=6)
        columns = ("date", "student", "class", "exam", "lesson", "score", "wpm", "source")
        self.results_tree = ttk.Treeview(tab, columns=columns, show="headings", selectmode="browse")
        for column, label, width in [("date", "Date", 145), ("student", "Student", 160), ("class", "Class", 90), ("exam", "Exam / session", 180), ("lesson", "Passage", 220), ("score", "Score", 78), ("wpm", "WPM", 64), ("source", "Source", 100)]:
            self.results_tree.heading(column, text=label)
            self.results_tree.column(column, width=width, anchor="w")
        self.results_tree.grid(row=2, column=0, sticky="nsew")
        self.results_tree.bind("<Double-1>", lambda _e: self.view_result())
        scroll = ttk.Scrollbar(tab, command=self.results_tree.yview)
        scroll.grid(row=2, column=1, sticky="ns")
        self.results_tree.configure(yscrollcommand=scroll.set)

    def refresh(self):
        self.refresh_lessons()
        self.refresh_students()
        self.refresh_results()
        lessons = self.db.list_lessons()
        students = self.db.list_students()
        attempts = self.db.list_attempts()
        average = sum(float(item.get("overall_score", 0)) for item in attempts) / len(attempts) if attempts else 0
        if hasattr(self, "metric_labels"):
            self.metric_labels["dictations"].configure(text=str(len(lessons)))
            self.metric_labels["students"].configure(text=str(len(students)))
            self.metric_labels["attempts"].configure(text=str(len(attempts)))
            self.metric_labels["average"].configure(text=f"{average:.1f}%")

    def refresh_lessons(self):
        self.lessons_tree.delete(*self.lessons_tree.get_children())
        for lesson in self.db.list_lessons():
            updated = lesson.get("updated_at", "").replace("T", " ")[:16]
            mode = "Sentence" if lesson.get("sentence_mode", 1) else "Passage"
            self.lessons_tree.insert("", "end", iid=str(lesson["id"]), values=(lesson["title"], "Français" if lesson["language"] == "fr" else "English", mode, lesson.get("category", ""), lesson.get("difficulty", ""), word_count(lesson.get("text", "")), updated))

    def selected_id(self, tree: ttk.Treeview) -> int | None:
        selection = tree.selection()
        return int(selection[0]) if selection else None

    def edit_lesson(self):
        lesson_id = self.selected_id(self.lessons_tree)
        if lesson_id is None:
            messagebox.showinfo("Select a dictation", "Select a dictation to edit.", parent=self)
            return
        lesson = self.db.get_lesson(lesson_id)
        if lesson:
            LessonEditor(self.app, lesson)

    def duplicate_lesson(self):
        lesson_id = self.selected_id(self.lessons_tree)
        if lesson_id is None:
            return
        self.db.duplicate_lesson(lesson_id)
        self.app.refresh_all()

    def delete_lesson(self):
        lesson_id = self.selected_id(self.lessons_tree)
        if lesson_id is None:
            return
        if messagebox.askyesno("Delete dictation", "Delete this dictation? Existing results will remain available.", parent=self):
            self.db.delete_lesson(lesson_id)
            self.app.refresh_all()

    def export_lessons_json(self):
        path = filedialog.asksaveasfilename(parent=self, defaultextension=".json", filetypes=[("DictaType lessons", "*.json")], initialfile="dictatype_lessons.json")
        if not path:
            return
        payload = {"format": "DictaType lessons", "version": 1, "lessons": self.db.export_lessons()}
        Path(path).write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
        messagebox.showinfo("Export complete", "The lessons were exported.", parent=self)

    def import_lessons_json(self):
        path = filedialog.askopenfilename(parent=self, filetypes=[("DictaType lessons", "*.json")])
        if not path:
            return
        try:
            payload = json.loads(Path(path).read_text(encoding="utf-8"))
            lessons = payload.get("lessons", payload if isinstance(payload, list) else [])
            count = self.db.import_lessons(lessons)
            self.app.refresh_all()
            messagebox.showinfo("Import complete", f"Imported {count} dictation(s).", parent=self)
        except Exception as exc:
            messagebox.showerror("Import failed", str(exc), parent=self)

    def refresh_students(self):
        self.students_tree.delete(*self.students_tree.get_children())
        for student in self.db.list_students():
            created = student.get("created_at", "").replace("T", " ")[:16]
            self.students_tree.insert(
                "", "end", iid=str(student["id"]),
                values=(
                    student["name"],
                    student.get("class_name", ""),
                    student.get("identifier", "") or "Auto",
                    "Active" if student.get("active", 1) else "Disabled",
                    created,
                ),
            )

    def student_form(self, student: dict[str, Any] | None = None):
        dialog = tk.Toplevel(self.app)
        dialog.title("Edit student profile" if student else "Add student profile")
        dialog.geometry("520x360")
        dialog.resizable(False, False)
        dialog.transient(self.app)
        dialog.grab_set()
        dialog.configure(bg=self.app.colors["bg"])

        frame = ttk.Frame(dialog, padding=24)
        frame.pack(fill="both", expand=True)
        frame.columnconfigure(1, weight=1)
        ttk.Label(frame, text="Student profile", style="PageTitle.TLabel").grid(row=0, column=0, columnspan=2, sticky="w")
        ttk.Label(frame, text="Students do not need a password or PIN. They select their profile from the student sign-in screen.", style="Muted.TLabel", wraplength=450).grid(row=1, column=0, columnspan=2, sticky="w", pady=(3, 16))

        name_var = tk.StringVar(value=student.get("name", "") if student else "")
        class_var = tk.StringVar(value=student.get("class_name", "") if student else "")
        id_var = tk.StringVar(value=student.get("identifier", "") if student else "")
        active_var = tk.BooleanVar(value=bool(student.get("active", 1)) if student else True)

        fields = [("Name", name_var), ("Class", class_var), ("Profile ID", id_var)]
        for row, (label, variable) in enumerate(fields, start=2):
            ttk.Label(frame, text=label).grid(row=row, column=0, sticky="w", padx=(0, 14), pady=8)
            ttk.Entry(frame, textvariable=variable).grid(row=row, column=1, sticky="ew", pady=8)
        ttk.Label(frame, text="Leave Profile ID blank to generate one automatically.", style="Muted.TLabel").grid(row=5, column=1, sticky="w", pady=(0, 6))
        ttk.Checkbutton(frame, text="Show this profile on the student sign-in screen", variable=active_var).grid(row=6, column=1, sticky="w", pady=(4, 10))

        def save():
            name = name_var.get().strip()
            if not name:
                messagebox.showwarning("Missing name", "Enter the student's name.", parent=dialog)
                return
            try:
                self.db.save_student(name, class_var.get(), id_var.get(), student.get("id") if student else None, active=active_var.get())
            except ValueError as exc:
                messagebox.showerror("Could not save student", str(exc), parent=dialog)
                return
            dialog.destroy()
            self.app.refresh_all()
            self.app.login_page.refresh_profiles()

        actions = ttk.Frame(frame)
        actions.grid(row=7, column=0, columnspan=2, sticky="e", pady=(18, 0))
        RoundedButton(actions, text="Cancel", style="Secondary.TButton", command=dialog.destroy).pack(side="right")
        RoundedButton(actions, text="Save profile", style="Accent.TButton", command=save).pack(side="right", padx=(0, 8))

    def add_student(self):
        self.student_form()

    def edit_student(self):
        student_id = self.selected_id(self.students_tree)
        if student_id is None:
            return
        student = self.db.get_student(student_id)
        if student:
            self.student_form(student)

    def view_student_history(self):
        student_id = self.selected_id(self.students_tree)
        if student_id is None:
            messagebox.showinfo("Select a student", "Select a student to view their complete result history.", parent=self)
            return
        student = self.db.get_student(student_id)
        if student:
            StudentHistoryDialog(self.app, student)

    def delete_student(self):
        student_id = self.selected_id(self.students_tree)
        if student_id is None:
            return
        if messagebox.askyesno("Delete student", "Delete this profile? Their previous results will remain available in Results.", parent=self):
            self.db.delete_student(student_id)
            self.app.refresh_all()
            self.app.login_page.refresh_profiles()

    def filtered_attempts(self):
        query = self.result_search_var.get().strip().casefold()
        attempts = self.db.list_attempts()
        if not query:
            return attempts
        return [
            item for item in attempts
            if query in " ".join([
                item.get("student_name", ""), item.get("class_name", ""),
                item.get("lesson_title", ""), item.get("exam_title", ""),
                item.get("source", ""),
            ]).casefold()
        ]

    def refresh_results(self):
        self.results_tree.delete(*self.results_tree.get_children())
        attempts = self.filtered_attempts()
        for attempt in attempts:
            created = attempt.get("created_at", "").replace("T", " ")[:16]
            exam = attempt.get("exam_title", "") or "Practice"
            if attempt.get("exam_item_count", 0):
                exam += f" {attempt.get('exam_item_index', 0)}/{attempt.get('exam_item_count', 0)}"
            self.results_tree.insert(
                "", "end", iid=str(attempt["id"]),
                values=(created, attempt.get("student_name", ""), attempt.get("class_name", ""), exam, attempt.get("lesson_title", ""), f"{attempt.get('overall_score', 0):.1f}%", f"{attempt.get('wpm', 0):.1f}", attempt.get("source", "")),
            )
        self.result_count_label.configure(text=f"{len(attempts)} result(s)")

    def view_result(self):
        attempt_id = self.selected_id(self.results_tree)
        if attempt_id is None:
            return
        attempt = self.db.get_attempt(attempt_id)
        if attempt:
            ResultDialog(self.app, attempt)

    def view_result_student_history(self):
        attempt_id = self.selected_id(self.results_tree)
        if attempt_id is None:
            messagebox.showinfo("Select a result", "Select one result to identify the student.", parent=self)
            return
        attempt = self.db.get_attempt(attempt_id)
        if not attempt:
            return
        student = self.db.get_student(attempt.get("student_id")) if attempt.get("student_id") else None
        if student is None:
            student = {
                "id": None,
                "name": attempt.get("student_name", "Student"),
                "class_name": attempt.get("class_name", ""),
                "identifier": "",
            }
        StudentHistoryDialog(self.app, student)

    def save_selected_exam_report(self):
        attempt_id = self.selected_id(self.results_tree)
        if attempt_id is None:
            messagebox.showinfo("Select an exam result", "Select any passage from the student's exam first.", parent=self)
            return
        attempt = self.db.get_attempt(attempt_id)
        if not attempt or not str(attempt.get("exam_session_id", "")).strip() or not str(attempt.get("source", "")).startswith("classroom-exam"):
            messagebox.showinfo("Not an exam result", "That result is not part of an Exam session. Full exam reports combine every passage from one exam.", parent=self)
            return
        items = self.db.list_exam_attempts(
            attempt.get("exam_session_id", ""),
            attempt.get("student_id"),
            attempt.get("student_name", ""),
            attempt.get("class_name", ""),
        )
        if not items:
            messagebox.showerror("Exam report", "The exam passages could not be found.", parent=self)
            return
        safe_student = "".join(c if c.isalnum() or c in "-_ " else "_" for c in str(attempt.get("student_name", "Student"))).strip().replace(" ", "_") or "Student"
        safe_exam = "".join(c if c.isalnum() or c in "-_ " else "_" for c in str(attempt.get("exam_title", "Exam"))).strip().replace(" ", "_") or "Exam"
        path = filedialog.asksaveasfilename(
            parent=self,
            defaultextension=".pdf",
            filetypes=[("PDF document", "*.pdf")],
            initialfile=f"DictaType_{safe_student}_{safe_exam}.pdf",
        )
        if not path:
            return
        try:
            save_exam_pdf(items, path)
            messagebox.showinfo("Exam report saved", f"One PDF containing all {len(items)} passage(s) was created.", parent=self)
        except Exception as exc:
            messagebox.showerror("Exam report failed", str(exc), parent=self)

    def delete_result(self):
        attempt_id = self.selected_id(self.results_tree)
        if attempt_id is not None and messagebox.askyesno("Delete result", "Permanently delete this result?", parent=self):
            self.db.delete_attempt(attempt_id)
            self.app.refresh_all()

    def export_results_csv(self):
        path = filedialog.asksaveasfilename(parent=self, defaultextension=".csv", filetypes=[("CSV file", "*.csv")], initialfile="dictatype_results.csv")
        if not path:
            return
        rows = self.filtered_attempts()
        with open(path, "w", newline="", encoding="utf-8-sig") as handle:
            writer = csv.writer(handle)
            writer.writerow(["Date", "Student", "Class", "Exam / Session", "Exam Item", "Dictation / Passage", "Overall Score", "Word Accuracy", "Character Accuracy", "WPM", "Duration Seconds", "Replays", "Source", "Teacher Comment"])
            for item in rows:
                writer.writerow([item.get("created_at"), item.get("student_name"), item.get("class_name"), item.get("exam_title", ""), f"{item.get('exam_item_index',0)}/{item.get('exam_item_count',0)}" if item.get("exam_item_count",0) else "", item.get("lesson_title"), item.get("overall_score"), item.get("score_word"), item.get("score_char"), item.get("wpm"), item.get("duration_seconds"), item.get("replay_count"), item.get("source"), item.get("teacher_comment")])
        messagebox.showinfo("Export complete", "The CSV file was created.", parent=self)

    def export_results_xlsx(self):
        path = filedialog.asksaveasfilename(parent=self, defaultextension=".xlsx", filetypes=[("Excel workbook", "*.xlsx")], initialfile="dictatype_results.xlsx")
        if not path:
            return
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.utils import get_column_letter

            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Results"
            headers = ["Date", "Student", "Class", "Exam / Session", "Exam Item", "Dictation / Passage", "Overall Score (%)", "Word Accuracy (%)", "Character Accuracy (%)", "WPM", "Duration (seconds)", "Replays", "Source", "Teacher Comment"]
            sheet.append(headers)
            for item in self.filtered_attempts():
                sheet.append([item.get("created_at"), item.get("student_name"), item.get("class_name"), item.get("exam_title", ""), f"{item.get('exam_item_index',0)}/{item.get('exam_item_count',0)}" if item.get("exam_item_count",0) else "", item.get("lesson_title"), item.get("overall_score"), item.get("score_word"), item.get("score_char"), item.get("wpm"), item.get("duration_seconds"), item.get("replay_count"), item.get("source"), item.get("teacher_comment")])
            for cell in sheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="496FF2")
                cell.alignment = Alignment(horizontal="center")
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions
            widths = [22, 22, 15, 26, 10, 35, 18, 18, 20, 10, 20, 10, 14, 35]
            for index, width in enumerate(widths, start=1):
                sheet.column_dimensions[get_column_letter(index)].width = width
            workbook.save(path)
            messagebox.showinfo("Export complete", "The Excel workbook was created.", parent=self)
        except Exception as exc:
            messagebox.showerror("Export failed", str(exc), parent=self)


class ClassroomPage(ttk.Frame):
    def __init__(self, app: "DictaTypeApp", master):
        super().__init__(master, padding=24, style="Page.TFrame")
        self.app = app
        self.db = app.db
        self.lesson_ids: list[int] = []
        self._build()

    def _build(self):
        ttk.Label(self, text="Local classroom & exams", style="PageTitle.TLabel").pack(anchor="w")
        ttk.Label(
            self,
            text="Choose Classroom for normal practice or Exam for a controlled multi-passage assessment. Students join from a browser on the same local network.",
            style="Muted.TLabel",
            wraplength=900,
        ).pack(anchor="w", pady=(4, 16))

        card = Card(self)
        card.pack(fill="x")
        card.columnconfigure(1, weight=1)

        ttk.Label(card, text="Session type").grid(row=0, column=0, sticky="w", padx=(0, 12), pady=7)
        self.session_type_var = tk.StringVar(value="Classroom")
        self.session_type_combo = ttk.Combobox(
            card,
            textvariable=self.session_type_var,
            values=["Classroom", "Exam"],
            state="readonly",
            width=18,
        )
        self.session_type_combo.grid(row=0, column=1, sticky="w", pady=7)
        self.session_type_combo.bind("<<ComboboxSelected>>", lambda _e: self._session_type_changed())

        ttk.Label(card, text="Starting port").grid(row=0, column=2, sticky="w", padx=(18, 8), pady=7)
        self.port_var = tk.IntVar(value=8765)
        ttk.Spinbox(card, from_=1024, to=65535, textvariable=self.port_var, width=10).grid(row=0, column=3, sticky="w", pady=7)

        ttk.Label(card, text="Session title").grid(row=1, column=0, sticky="w", padx=(0, 12), pady=7)
        self.exam_title_var = tk.StringVar(value="Classroom Practice")
        ttk.Entry(card, textvariable=self.exam_title_var).grid(row=1, column=1, columnspan=3, sticky="ew", pady=7)

        self.session_help_var = tk.StringVar()
        ttk.Label(card, textvariable=self.session_help_var, style="CardMuted.TLabel", wraplength=780).grid(
            row=2, column=1, columnspan=3, sticky="w", pady=(0, 8)
        )

        ttk.Label(card, text="Passages / dictations").grid(row=3, column=0, sticky="nw", padx=(0, 12), pady=(10, 7))
        list_frame = ttk.Frame(card, style="Field.TFrame")
        list_frame.grid(row=3, column=1, columnspan=3, sticky="nsew", pady=(10, 7))
        list_frame.columnconfigure(0, weight=1)
        self.lesson_listbox = tk.Listbox(
            list_frame,
            selectmode="browse",
            height=7,
            exportselection=False,
            activestyle="none",
            relief="flat",
            borderwidth=0,
            font=("Segoe UI", 10),
            bg=self.app.colors["field"],
            fg=self.app.colors["text"],
            selectbackground=self.app.colors["accent"],
            selectforeground=self.app.colors["text"],
        )
        self.lesson_listbox.grid(row=0, column=0, sticky="ew")
        scroll = ttk.Scrollbar(list_frame, command=self.lesson_listbox.yview)
        scroll.grid(row=0, column=1, sticky="ns")
        self.lesson_listbox.configure(yscrollcommand=scroll.set)

        selection_buttons = ttk.Frame(card, style="Card.TFrame")
        selection_buttons.grid(row=4, column=1, columnspan=3, sticky="w", pady=(0, 8))
        self.select_all_button = RoundedButton(selection_buttons, text="Select all passages", style="Secondary.TButton", command=self.select_all)
        self.select_all_button.pack(side="left")
        RoundedButton(selection_buttons, text="Clear selection", style="Secondary.TButton", command=self.clear_selection).pack(side="left", padx=8)

        self.allow_new_profiles_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            card,
            text="Allow new students to create a profile when they join",
            variable=self.allow_new_profiles_var,
        ).grid(row=5, column=1, columnspan=3, sticky="w", pady=(2, 0))

        self.enhanced_audio_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            card,
            text="Use teacher-computer audio for clearer and consistent dictation",
            variable=self.enhanced_audio_var,
        ).grid(row=6, column=1, columnspan=3, sticky="w", pady=(8, 0))

        self.builtin_french_var = tk.BooleanVar(value=True)
        self.builtin_french_check = ttk.Checkbutton(
            card,
            text="Use DictaType built-in French neural voice (recommended)",
            variable=self.builtin_french_var,
        )
        self.builtin_french_check.grid(row=7, column=1, columnspan=3, sticky="w", pady=(3, 0))

        self.french_clarity_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            card,
            text="French dictation pace (slower, clearer delivery)",
            variable=self.french_clarity_var,
        ).grid(row=8, column=1, columnspan=3, sticky="w", pady=(3, 0))
        self.french_voice_status_var = tk.StringVar()
        ttk.Label(
            card,
            textvariable=self.french_voice_status_var,
            style="CardMuted.TLabel",
            wraplength=780,
        ).grid(row=9, column=1, columnspan=3, sticky="w", pady=(2, 8))
        self._refresh_french_voice_status()

        self.performance_status_var = tk.StringVar()
        ttk.Label(
            card,
            textvariable=self.performance_status_var,
            style="CardMuted.TLabel",
            wraplength=780,
        ).grid(row=10, column=1, columnspan=3, sticky="w", pady=(3, 5))
        self._refresh_performance_status()

        buttons = ttk.Frame(card, style="Card.TFrame")
        buttons.grid(row=11, column=0, columnspan=4, sticky="ew", pady=(12, 0))
        self.start_button = RoundedButton(buttons, text="Start classroom", style="Accent.TButton", command=self.start_server)
        self.start_button.pack(side="left")
        self.stop_button = RoundedButton(buttons, text="Stop", style="Danger.TButton", command=self.stop_server, state="disabled")
        self.stop_button.pack(side="left", padx=8)

        session_card = Card(self)
        session_card.pack(fill="x", pady=14)
        session_card.columnconfigure(1, weight=1)
        ttk.Label(session_card, text="Student address", style="Muted.TLabel").grid(row=0, column=0, sticky="w", padx=(0, 14), pady=5)
        self.url_var = tk.StringVar(value="Classroom is not running")
        ttk.Entry(session_card, textvariable=self.url_var, state="readonly").grid(row=0, column=1, sticky="ew", pady=5)
        RoundedButton(session_card, text="Open", style="Secondary.TButton", command=lambda: webbrowser.open(self.url_var.get())).grid(row=0, column=2, padx=(8, 0))
        ttk.Label(session_card, text="Session code", style="Muted.TLabel").grid(row=1, column=0, sticky="w", padx=(0, 14), pady=5)
        self.code_var = tk.StringVar(value="------")
        ttk.Label(session_card, textvariable=self.code_var, style="Code.TLabel").grid(row=1, column=1, sticky="w", pady=5)
        RoundedButton(session_card, text="Copy details", style="Secondary.TButton", command=self.copy_details).grid(row=1, column=2, padx=(8, 0))
        self.session_summary_var = tk.StringVar(value="")
        ttk.Label(session_card, textvariable=self.session_summary_var, style="Muted.TLabel").grid(row=2, column=1, columnspan=2, sticky="w", pady=(3, 0))

        recent_header = ttk.Frame(self, style="Page.TFrame")
        recent_header.pack(fill="x", pady=(10, 8))
        ttk.Label(recent_header, text="Recent classroom submissions", style="SectionTitle.TLabel").pack(side="left")
        RoundedButton(recent_header, text="View result", style="Secondary.TButton", command=self.view_submission).pack(side="right")
        RoundedButton(recent_header, text="Save full exam PDF", style="Accent.TButton", command=self.save_exam_report).pack(side="right", padx=8)

        frame = ttk.Frame(self)
        frame.pack(fill="both", expand=True)
        frame.columnconfigure(0, weight=1)
        frame.rowconfigure(0, weight=1)
        columns = ("date", "student", "class", "type", "exam", "lesson", "score")
        self.tree = ttk.Treeview(frame, columns=columns, show="headings", selectmode="browse")
        for column, title, width in [
            ("date", "Date", 135), ("student", "Student", 160), ("class", "Class", 90),
            ("type", "Type", 85), ("exam", "Exam / session", 175), ("lesson", "Passage", 220), ("score", "Score", 75),
        ]:
            self.tree.heading(column, text=title)
            self.tree.column(column, width=width, anchor="w")
        self.tree.grid(row=0, column=0, sticky="nsew")
        self.tree.bind("<Double-1>", self.view_submission)
        scroll2 = ttk.Scrollbar(frame, command=self.tree.yview)
        scroll2.grid(row=0, column=1, sticky="ns")
        self.tree.configure(yscrollcommand=scroll2.set)
        self._session_type_changed()

    def _refresh_french_voice_status(self):
        if builtin_french_available():
            self.french_voice_status_var.set(
                "Built-in fr-FR neural voice is ready. French passages are generated locally on the teacher computer and sent to every student, so pronunciation does not depend on the student's browser or Windows language pack."
            )
            self.builtin_french_check.configure(state="normal")
        else:
            diag = french_voice_diagnostics(synthesize=False)
            reason = str(diag.get("reason") or "Built-in French neural voice is unavailable.")
            self.french_voice_status_var.set(
                f"{reason} DictaType will use an installed Windows/browser French voice as a fallback."
            )
            self.builtin_french_check.configure(state="disabled")
            self.builtin_french_var.set(False)

    def _refresh_performance_status(self):
        profile = self.app.performance
        policy = (
            "Exam audio is pre-generated to disk and the neural model is released before students join."
            if profile.low_memory
            else "Exam audio is pre-generated; normal classroom audio may be generated on first use."
        )
        self.performance_status_var.set(
            f"Performance: {profile.label} · {profile.hardware_summary}. {policy}"
        )

    def _session_type_changed(self):
        is_exam = self.session_type_var.get() == "Exam"
        self.lesson_listbox.configure(selectmode="extended" if is_exam else "browse")
        self.select_all_button.configure(state="normal" if is_exam else "disabled")
        if is_exam:
            self.session_help_var.set("Exam mode can contain several passages. Every passage is saved under one exam session and can later be exported as one combined correction PDF.")
            if self.exam_title_var.get() in {"", "Classroom Practice"}:
                self.exam_title_var.set("DictaType Exam")
            self.allow_new_profiles_var.set(False)
            self.start_button.configure(text="Start exam")
        else:
            selections = list(self.lesson_listbox.curselection())
            if len(selections) > 1:
                keep = selections[0]
                self.lesson_listbox.selection_clear(0, "end")
                self.lesson_listbox.selection_set(keep)
            self.session_help_var.set("Classroom mode distributes one dictation at a time for normal teaching or practice.")
            if self.exam_title_var.get() in {"", "DictaType Exam"}:
                self.exam_title_var.set("Classroom Practice")
            self.allow_new_profiles_var.set(True)
            self.start_button.configure(text="Start classroom")

    def select_all(self):
        if self.session_type_var.get() == "Exam":
            self.lesson_listbox.selection_set(0, "end")

    def clear_selection(self):
        self.lesson_listbox.selection_clear(0, "end")

    def refresh(self):
        self._refresh_french_voice_status()
        lessons = self.db.list_lessons()
        selected_ids = {self.lesson_ids[index] for index in self.lesson_listbox.curselection() if index < len(self.lesson_ids)} if self.lesson_ids else set()
        self.lesson_ids = []
        self.lesson_listbox.delete(0, "end")
        for lesson in lessons:
            mode = "Sentence" if lesson.get("sentence_mode", 1) else "Passage"
            label = f"{lesson['title']}  ·  {'FR' if lesson['language'] == 'fr' else 'EN'}  ·  {mode}  ·  {word_count(lesson.get('text',''))} words"
            self.lesson_ids.append(int(lesson["id"]))
            self.lesson_listbox.insert("end", label)
            if lesson["id"] in selected_ids:
                self.lesson_listbox.selection_set(len(self.lesson_ids) - 1)
        if self.lesson_ids and not self.lesson_listbox.curselection():
            self.lesson_listbox.selection_set(0)
        self._session_type_changed()

        self.tree.delete(*self.tree.get_children())
        limit = self.app.performance.classroom_result_limit
        for attempt in [item for item in self.db.list_attempts(limit) if str(item.get("source", "")).startswith("classroom")]:
            is_exam = str(attempt.get("source", "")).startswith("classroom-exam")
            exam = attempt.get("exam_title", "") or ("Exam" if is_exam else "Classroom")
            if is_exam and attempt.get("exam_item_count", 0) > 1:
                exam += f" {attempt.get('exam_item_index', 0)}/{attempt.get('exam_item_count', 0)}"
            self.tree.insert(
                "", "end", iid=str(attempt["id"]),
                values=(attempt.get("created_at", "").replace("T", " ")[:16], attempt.get("student_name", ""), attempt.get("class_name", ""), "Exam" if is_exam else "Classroom", exam, attempt.get("lesson_title", ""), f"{attempt.get('overall_score',0):.1f}%"),
            )

    def start_server(self):
        indices = list(self.lesson_listbox.curselection())
        lessons = []
        for index in indices:
            if index < len(self.lesson_ids):
                lesson = self.db.get_lesson(self.lesson_ids[index])
                if lesson:
                    lessons.append(lesson)
        if not lessons:
            messagebox.showwarning("Select a dictation", "Select at least one dictation or passage to distribute.", parent=self)
            return

        is_exam = self.session_type_var.get() == "Exam"
        if not is_exam and len(lessons) != 1:
            messagebox.showwarning("Classroom mode", "Classroom mode uses one dictation at a time. Choose Exam when you want multiple passages.", parent=self)
            return
        title = self.exam_title_var.get().strip()
        if is_exam and not title:
            messagebox.showwarning("Exam title", "Give the exam a title so all of its passages can be grouped together in Results.", parent=self)
            return
        if not title:
            title = lessons[0].get("title", "Classroom Dictation")

        def audio_progress(done: int, total: int, passage_title: str) -> None:
            if total:
                self.session_summary_var.set(
                    f"Preparing audio {done}/{total} · {passage_title}"
                )
                self.update_idletasks()

        try:
            self.start_button.configure(state="disabled")
            url, code = self.app.classroom_server.start(
                lessons,
                self.port_var.get(),
                exam_title=title,
                allow_new_profiles=self.allow_new_profiles_var.get(),
                session_type="exam" if is_exam else "classroom",
                enhanced_audio=self.enhanced_audio_var.get(),
                french_clarity=self.french_clarity_var.get(),
                builtin_french=self.builtin_french_var.get(),
                performance_mode=self.app.performance.requested_mode,
                precache_audio=is_exam or self.app.performance.low_memory,
                progress_callback=audio_progress,
            )
            self.url_var.set(url)
            self.code_var.set(code)
            mode_text = "Exam" if is_exam else "Classroom"
            prepared = self.app.classroom_server.audio_prepared_count
            total_audio = self.app.classroom_server.audio_prepared_total
            audio_text = f" · audio ready {prepared}/{total_audio}" if total_audio else ""
            self.session_summary_var.set(
                f"{mode_text}: {title} · {len(lessons)} passage(s){audio_text} · {self.app.performance.label}"
            )
            self.start_button.configure(state="disabled")
            self.stop_button.configure(state="normal")
            self.lesson_listbox.configure(state="disabled")
            self.session_type_combo.configure(state="disabled")
            messagebox.showinfo(
                f"{mode_text} started",
                f"Students can open:\n{url}\n\nSession code: {code}\n\nPassages: {len(lessons)}",
                parent=self,
            )
        except Exception as exc:
            self.start_button.configure(state="normal")
            self.session_summary_var.set("")
            messagebox.showerror(f"Could not start {'exam' if is_exam else 'classroom'}", str(exc), parent=self)

    def stop_server(self):
        self.app.classroom_server.stop()
        self.url_var.set("Classroom is not running")
        self.code_var.set("------")
        self.session_summary_var.set("")
        self.start_button.configure(state="normal")
        self.stop_button.configure(state="disabled")
        self.lesson_listbox.configure(state="normal")
        self.session_type_combo.configure(state="readonly")
        self._session_type_changed()

    def copy_details(self):
        if not self.app.classroom_server.running:
            return
        mode = "Exam" if self.app.classroom_server.session_type == "exam" else "Classroom"
        text = f"DictaType {mode}\nAddress: {self.url_var.get()}\nSession code: {self.code_var.get()}\n{self.session_summary_var.get()}"
        self.clipboard_clear()
        self.clipboard_append(text)

    def selected_attempt(self) -> dict[str, Any] | None:
        selection = self.tree.selection()
        if not selection:
            return None
        return self.db.get_attempt(int(selection[0]))

    def view_submission(self, _event=None):
        attempt = self.selected_attempt()
        if attempt:
            ResultDialog(self.app, attempt)

    def save_exam_report(self):
        attempt = self.selected_attempt()
        if not attempt:
            messagebox.showinfo("Select an exam result", "Select any passage from a student's exam first.", parent=self)
            return
        if not str(attempt.get("source", "")).startswith("classroom-exam") or not str(attempt.get("exam_session_id", "")).strip():
            messagebox.showinfo("Not an exam result", "Select an Exam result. Classroom practice contains only an individual result report.", parent=self)
            return
        items = self.db.list_exam_attempts(
            attempt.get("exam_session_id", ""),
            attempt.get("student_id"),
            attempt.get("student_name", ""),
            attempt.get("class_name", ""),
        )
        if not items:
            messagebox.showerror("Exam report", "The exam passages could not be found.", parent=self)
            return
        safe_student = "".join(c if c.isalnum() or c in "-_ " else "_" for c in str(attempt.get("student_name", "Student"))).strip().replace(" ", "_") or "Student"
        safe_exam = "".join(c if c.isalnum() or c in "-_ " else "_" for c in str(attempt.get("exam_title", "Exam"))).strip().replace(" ", "_") or "Exam"
        path = filedialog.asksaveasfilename(
            parent=self,
            defaultextension=".pdf",
            filetypes=[("PDF document", "*.pdf")],
            initialfile=f"DictaType_{safe_student}_{safe_exam}.pdf",
        )
        if not path:
            return
        try:
            save_exam_pdf(items, path)
            messagebox.showinfo("Exam report saved", f"All {len(items)} passage(s) were saved in one PDF for correction.", parent=self)
        except Exception as exc:
            messagebox.showerror("Exam report failed", str(exc), parent=self)


class SettingsPage(ttk.Frame):
    def __init__(self, app: "DictaTypeApp", master):
        super().__init__(master, padding=24, style="Page.TFrame")
        self.app = app
        self.db = app.db
        self._build()

    def _build(self):
        header = ttk.Frame(self, style="Page.TFrame")
        header.pack(fill="x")
        ttk.Label(header, text="Settings & security", style="PageTitle.TLabel").pack(side="left")
        ttk.Label(header, text="TEACHER ONLY", style="RoleBadge.TLabel").pack(side="right")
        ttk.Label(
            self,
            text="Security controls, performance, local backups and system voice management.",
            style="Muted.TLabel",
        ).pack(anchor="w", pady=(4, 16))

        security = Card(self)
        security.pack(fill="x", pady=(0, 12))
        ttk.Label(security, text="Teacher security", style="CardSectionTitle.TLabel").grid(row=0, column=0, columnspan=3, sticky="w", pady=(0, 8))
        ttk.Label(
            security,
            text="Teacher features are available only after teacher authentication. Student sessions never load these screens.",
            style="CardMuted.TLabel",
            wraplength=760,
        ).grid(row=1, column=0, columnspan=3, sticky="w")
        RoundedButton(security, text="Change teacher PIN", style="Secondary.TButton", command=self.change_pin).grid(row=2, column=0, sticky="w", pady=(14, 0))
        RoundedButton(security, text="Lock now", style="Secondary.TButton", command=self.app.logout).grid(row=2, column=1, sticky="w", padx=(8, 0), pady=(14, 0))

        session = Card(self)
        session.pack(fill="x", pady=(0, 12))
        session.columnconfigure(1, weight=1)
        ttk.Label(session, text="Automatic teacher lock", style="CardSectionTitle.TLabel").grid(row=0, column=0, columnspan=3, sticky="w", pady=(0, 8))
        ttk.Label(session, text="Lock the teacher interface after inactivity.", style="CardMuted.TLabel").grid(row=1, column=0, sticky="w", padx=(0, 18))
        current = self.db.get_setting("teacher_auto_lock_minutes", "15")
        self.lock_var = tk.StringVar(value="Off" if current == "0" else f"{current} minutes")
        ttk.Combobox(
            session,
            textvariable=self.lock_var,
            values=["5 minutes", "10 minutes", "15 minutes", "30 minutes", "60 minutes", "Off"],
            state="readonly",
            width=16,
        ).grid(row=1, column=1, sticky="w")
        RoundedButton(session, text="Save", style="Accent.TButton", command=self.save_lock_setting).grid(row=1, column=2, sticky="e", padx=(12, 0))

        performance = Card(self)
        performance.pack(fill="x", pady=(0, 12))
        performance.columnconfigure(1, weight=1)
        ttk.Label(performance, text="Performance", style="CardSectionTitle.TLabel").grid(row=0, column=0, columnspan=3, sticky="w", pady=(0, 8))
        ttk.Label(
            performance,
            text="Automatic is recommended. It detects low-memory PCs and uses disk-cached exam audio, lower UI overhead and conservative neural-voice CPU settings.",
            style="CardMuted.TLabel",
            wraplength=760,
        ).grid(row=1, column=0, columnspan=3, sticky="w")
        saved_mode = self.db.get_setting("performance_mode", "auto")
        mode_labels = {"auto": "Automatic (recommended)", "low": "Low-memory / HDD", "standard": "Standard / SSD"}
        self.performance_var = tk.StringVar(value=mode_labels.get(saved_mode, mode_labels["auto"]))
        ttk.Label(performance, text="Mode").grid(row=2, column=0, sticky="w", pady=(12, 0), padx=(0, 12))
        ttk.Combobox(
            performance,
            textvariable=self.performance_var,
            values=list(mode_labels.values()),
            state="readonly",
            width=25,
        ).grid(row=2, column=1, sticky="w", pady=(12, 0))
        RoundedButton(performance, text="Apply", style="Accent.TButton", command=self.save_performance_setting).grid(row=2, column=2, sticky="e", pady=(12, 0), padx=(12, 0))
        self.performance_info = ttk.Label(performance, text="", style="CardMuted.TLabel")
        self.performance_info.grid(row=3, column=0, columnspan=3, sticky="w", pady=(8, 0))

        data = Card(self)
        data.pack(fill="x", pady=(0, 12))
        ttk.Label(data, text="Local data", style="CardSectionTitle.TLabel").pack(anchor="w", pady=(0, 8))
        ttk.Label(data, text=f"Data folder: {app_data_dir()}", style="CardMuted.TLabel").pack(anchor="w")
        buttons = ttk.Frame(data, style="Card.TFrame")
        buttons.pack(fill="x", pady=(12, 0))
        RoundedButton(buttons, text="Back up database", style="Accent.TButton", command=self.backup).pack(side="left")
        RoundedButton(buttons, text="Restore database", style="Secondary.TButton", command=self.restore).pack(side="left", padx=8)
        RoundedButton(buttons, text="Open data folder", style="Secondary.TButton", command=lambda: open_folder(app_data_dir())).pack(side="left")

        voices = Card(self)
        voices.pack(fill="x")
        ttk.Label(voices, text="System voices", style="CardSectionTitle.TLabel").pack(anchor="w", pady=(0, 8))
        self.voice_label = ttk.Label(voices, text="", style="CardMuted.TLabel")
        self.voice_label.pack(anchor="w")
        voice_buttons = ttk.Frame(voices, style="Card.TFrame")
        voice_buttons.pack(anchor="w", pady=(12, 0))
        RoundedButton(voice_buttons, text="Refresh installed voices", style="Secondary.TButton", command=self.refresh_voices).pack(side="left")
        RoundedButton(voice_buttons, text="Test French neural voice", style="Accent.TButton", command=self.test_french_voice).pack(side="left", padx=(8, 0))
        self.refresh()

    def refresh(self):
        if hasattr(self, "performance_info"):
            self.performance_info.configure(
                text=f"Active: {self.app.performance.label} · {self.app.performance.hardware_summary}"
            )
        english = len(self.app.voices_for_language("en"))
        french = len(self.app.voices_for_language("fr"))
        neural = "Built-in French neural voice ready" if builtin_french_available() else "Built-in French neural voice unavailable"
        system_count = sum(1 for voice in self.app.voices if not voice.id.startswith("dictatype:piper:"))
        french_system = sum(
            1
            for voice in self.app.voices_for_language("fr")
            if not voice.id.startswith("dictatype:piper:")
        )
        self.voice_label.configure(
            text=(
                f"{neural}. Windows voices: {system_count} total · "
                f"{english} English-compatible · {french_system} French-compatible."
            )
        )

    def save_performance_setting(self):
        reverse = {
            "Automatic (recommended)": "auto",
            "Low-memory / HDD": "low",
            "Standard / SSD": "standard",
        }
        requested = reverse.get(self.performance_var.get(), "auto")
        self.db.set_setting("performance_mode", requested)
        self.app.performance = resolve_performance_profile(requested)
        apply_runtime_hints(self.app.performance)
        self.refresh()
        classroom = self.app.pages.get("classroom")
        if classroom is not None and hasattr(classroom, "_refresh_performance_status"):
            classroom._refresh_performance_status()
        messagebox.showinfo(
            "Performance updated",
            f"DictaType is now using {self.app.performance.label}. The setting is saved for future starts.",
            parent=self,
        )

    def save_lock_setting(self):
        value = self.lock_var.get()
        minutes = "0" if value == "Off" else value.split()[0]
        self.db.set_setting("teacher_auto_lock_minutes", minutes)
        self.app.touch_activity()
        messagebox.showinfo("Security updated", "Automatic teacher locking was updated.", parent=self)

    def change_pin(self):
        current = simpledialog.askstring("Current PIN", "Enter the current teacher PIN:", show="●", parent=self)
        if current is None:
            return
        if not self.db.check_pin(current):
            messagebox.showerror("Incorrect PIN", "The current PIN is incorrect.", parent=self)
            return
        new = simpledialog.askstring("New PIN", "Enter a new 6 to 12 digit teacher PIN:", show="●", parent=self)
        if not new or not new.isdigit() or not 6 <= len(new) <= 12:
            messagebox.showwarning("Invalid PIN", "Use a teacher PIN containing 6 to 12 digits.", parent=self)
            return
        confirm = simpledialog.askstring("Confirm PIN", "Enter the new PIN again:", show="●", parent=self)
        if new != confirm:
            messagebox.showerror("PIN mismatch", "The PINs do not match.", parent=self)
            return
        self.db.set_pin(new)
        self.db.set_setting("security_setup_complete", "1")
        messagebox.showinfo("PIN changed", "The teacher PIN was updated.", parent=self)

    def backup(self):
        default = f"DictaType_backup_{datetime.now():%Y-%m-%d}.db"
        path = filedialog.asksaveasfilename(parent=self, defaultextension=".db", filetypes=[("DictaType database", "*.db")], initialfile=default)
        if not path:
            return
        try:
            self.db.backup(Path(path))
            messagebox.showinfo("Backup complete", "The database backup was created.", parent=self)
        except Exception as exc:
            messagebox.showerror("Backup failed", str(exc), parent=self)

    def restore(self):
        path = filedialog.askopenfilename(parent=self, filetypes=[("DictaType database", "*.db"), ("All files", "*.*")])
        if not path:
            return
        if not messagebox.askyesno("Restore database", "This replaces the current local database. Continue?", parent=self):
            return
        try:
            self.app.classroom_server.stop()
            self.db.restore(Path(path))
            self.app.refresh_all()
            messagebox.showinfo("Restore complete", "The database was restored.", parent=self)
        except Exception as exc:
            messagebox.showerror("Restore failed", str(exc), parent=self)

    def test_french_voice(self):
        diag = french_voice_diagnostics(synthesize=False)
        if not diag.get("model_exists") or not diag.get("config_exists"):
            searched = "\n".join(str(item) for item in diag.get("searched_voice_directories", []))
            messagebox.showerror(
                "French neural voice missing",
                "DictaType could not find the French voice files.\n\nSearched locations:\n" + searched,
                parent=self,
            )
            return

        def on_error(exc):
            self.after(0, lambda: messagebox.showerror("French voice test failed", str(exc), parent=self))

        self.app.speech.speak(
            "Bonjour. Ceci est un test de la voix française de DictaType.",
            voice_id=BUNDLED_FRENCH_VOICE_ID,
            rate=145,
            on_error=on_error,
        )

    def refresh_voices(self):
        self.app.refresh_voices()
        self.refresh()
        messagebox.showinfo("Voices refreshed", "Installed system voices were scanned again.", parent=self)


class LoginPage(ttk.Frame):
    def __init__(self, app: "DictaTypeApp", master):
        super().__init__(master, style="Page.TFrame", padding=34)
        self.app = app
        self.role = "student"
        self._build()

    def _build(self):
        self.columnconfigure(0, weight=1)
        self.rowconfigure(0, weight=1)

        wrapper = ttk.Frame(self, style="Page.TFrame")
        wrapper.grid(row=0, column=0)
        wrapper.columnconfigure(0, weight=1)

        brand = ttk.Frame(wrapper, style="Page.TFrame")
        brand.grid(row=0, column=0, sticky="ew", pady=(0, 18))
        ttk.Label(brand, text="DictaType", style="LoginTitle.TLabel").pack()
        ttk.Label(brand, text="Bilingual dictation workspace", style="Muted.TLabel").pack(pady=(4, 0))

        card = Card(wrapper, padding=26)
        card.grid(row=1, column=0, sticky="ew")
        card.columnconfigure(0, weight=1)
        card.columnconfigure(1, weight=1)

        ttk.Label(card, text="Welcome", style="CardSectionTitle.TLabel").grid(row=0, column=0, columnspan=2, sticky="w")
        ttk.Label(card, text="Students use a simple profile. Teacher tools stay protected by the teacher PIN.", style="CardMuted.TLabel", wraplength=430).grid(row=1, column=0, columnspan=2, sticky="w", pady=(3, 14))

        role_bar = ttk.Frame(card, style="Card.TFrame")
        role_bar.grid(row=2, column=0, columnspan=2, sticky="ew", pady=(0, 16))
        self.student_role_button = RoundedButton(role_bar, text="Student", style="Accent.TButton", command=lambda: self.set_role("student"))
        self.student_role_button.pack(side="left", fill="x", expand=True)
        self.teacher_role_button = RoundedButton(role_bar, text="Teacher", style="Secondary.TButton", command=lambda: self.set_role("teacher"))
        self.teacher_role_button.pack(side="left", fill="x", expand=True, padx=(8, 0))

        self.student_area = ttk.Frame(card, style="Card.TFrame")
        self.student_area.grid(row=3, column=0, columnspan=2, sticky="ew")
        self.student_area.columnconfigure(0, weight=1)
        ttk.Label(self.student_area, text="Your name").grid(row=0, column=0, sticky="w", pady=(0, 6))
        self.student_name_var = tk.StringVar()
        self.student_name_entry = ttk.Entry(self.student_area, textvariable=self.student_name_var, width=42)
        self.student_name_entry.grid(row=1, column=0, sticky="ew", pady=(0, 10))
        ttk.Label(self.student_area, text="Class").grid(row=2, column=0, sticky="w", pady=(0, 6))
        self.student_class_var = tk.StringVar()
        self.student_class_entry = ttk.Entry(self.student_area, textvariable=self.student_class_var)
        self.student_class_entry.grid(row=3, column=0, sticky="ew", pady=(0, 10))
        self.student_class_entry.bind("<Return>", lambda _e: self.sign_in())
        student_actions = ttk.Frame(self.student_area, style="Card.TFrame")
        student_actions.grid(row=4, column=0, sticky="ew")
        RoundedButton(student_actions, text="Open my profile", style="Accent.TButton", command=self.sign_in).pack(side="left", fill="x", expand=True)
        RoundedButton(student_actions, text="Create my profile", style="Secondary.TButton", command=self.create_profile).pack(side="left", fill="x", expand=True, padx=(8, 0))
        ttk.Label(self.student_area, text="No password or student PIN is required. Use the same name and class every time so previous results stay together.", style="CardMuted.TLabel", wraplength=430).grid(row=5, column=0, sticky="w", pady=(10, 0))

        self.teacher_area = ttk.Frame(card, style="Card.TFrame")
        self.teacher_area.grid(row=4, column=0, columnspan=2, sticky="ew", pady=(2, 0))
        self.teacher_area.columnconfigure(0, weight=1)
        ttk.Label(self.teacher_area, text="Teacher PIN").grid(row=0, column=0, sticky="w", pady=(0, 6))
        self.pin_var = tk.StringVar()
        self.pin_entry = ttk.Entry(self.teacher_area, textvariable=self.pin_var, show="●")
        self.pin_entry.grid(row=1, column=0, sticky="ew")
        self.pin_entry.bind("<Return>", lambda _e: self.sign_in())
        self.teacher_signin_button = RoundedButton(self.teacher_area, text="Unlock teacher workspace", style="Accent.TButton", command=self.sign_in)
        self.teacher_signin_button.grid(row=2, column=0, sticky="ew", pady=(12, 0))

        self.error_label = ttk.Label(card, text="", style="ErrorCard.TLabel", wraplength=430)
        self.error_label.grid(row=5, column=0, columnspan=2, sticky="w", pady=(12, 0))

        ttk.Label(
            wrapper,
            text="Student sessions never load the Teacher Dashboard, Settings or classroom administration pages.",
            style="Muted.TLabel",
            wraplength=470,
            justify="center",
        ).grid(row=2, column=0, pady=(16, 0))
        self.set_role("student")

    def refresh_profiles(self):
        # Kept as a compatibility hook for teacher profile edits. No roster is shown to students.
        return

    def set_role(self, role: str):
        self.role = role
        self.error_label.configure(text="")
        self.pin_var.set("")
        if role == "student":
            self.teacher_area.grid_remove()
            self.student_area.grid()
            self.student_role_button._style_name = "Accent.TButton"
            self.teacher_role_button._style_name = "Secondary.TButton"
            self.student_role_button._draw()
            self.teacher_role_button._draw()
            self.student_name_entry.focus_set()
        else:
            self.student_area.grid_remove()
            self.teacher_area.grid()
            self.student_role_button._style_name = "Secondary.TButton"
            self.teacher_role_button._style_name = "Accent.TButton"
            self.student_role_button._draw()
            self.teacher_role_button._draw()
            self.pin_entry.focus_set()

    def create_profile(self):
        name = self.student_name_var.get().strip()
        class_name = self.student_class_var.get().strip()
        if not name:
            self.error_label.configure(text="Enter your name before creating a profile.")
            self.student_name_entry.focus_set()
            return
        try:
            student = self.app.db.create_student_profile(name, class_name)
        except ValueError as exc:
            self.error_label.configure(text=str(exc))
            return
        self.app.start_session("student", student)

    def sign_in(self):
        self.error_label.configure(text="")
        if self.role == "teacher":
            ok, message = self.app.login_teacher(self.pin_var.get())
        else:
            ok, message = self.app.login_student_by_name(self.student_name_var.get(), self.student_class_var.get())
        if not ok:
            self.error_label.configure(text=message)
            if self.role == "teacher":
                self.pin_var.set("")
                self.pin_entry.focus_set()
            else:
                self.student_name_entry.focus_set()

    def reset(self):
        self.pin_var.set("")
        self.error_label.configure(text="")
        self.set_role("student")


class DictaTypeApp(tk.Tk):
    def __init__(self, db_path: Path | None = None):
        super().__init__()
        self.db = Database(db_path)
        self.performance = resolve_performance_profile(self.db.get_setting("performance_mode", "auto"))
        apply_runtime_hints(self.performance)
        # The redesigned interface intentionally uses the grey + light-blue palette.
        self.colors = LIGHT
        self.speech = SpeechEngine()
        self.voices: list[Voice] = list_voices()
        self.current_role: str | None = None
        self.current_student: dict[str, Any] | None = None
        self.teacher_authenticated = False
        self.pages: dict[str, ttk.Frame] = {}
        self.nav_buttons: dict[str, RoundedButton] = {}
        self._auth_failures: dict[str, int] = {}
        self._auth_locked_until: dict[str, float] = {}
        self.last_activity = time.monotonic()
        self._ui_queue: queue.Queue[Callable[[], None]] = queue.Queue()
        self._closing = False
        self.classroom_server = ClassroomServer(self.db, on_submission=lambda: self.post(self.refresh_all))

        self.title(f"{APP_TITLE} {APP_VERSION}")
        self.geometry(self.db.get_setting("window_geometry", "1100x720"))
        self.minsize(900, 620)
        self.configure(bg=self.colors["bg"])
        self.option_add("*Font", ("Segoe UI", 10))
        self._configure_styles()
        self._build_root_views()
        self.protocol("WM_DELETE_WINDOW", self.on_close)
        self.bind("<Escape>", lambda _e: self.attributes("-fullscreen", False))
        self.bind_all("<KeyPress>", self.touch_activity, add="+")
        self.bind_all("<Button>", self.touch_activity, add="+")
        self.after(100 if self.performance.low_memory else 60, self._drain_ui_queue)
        self.after(30_000, self._security_tick)
        self.show_login()

    def post(self, callback: Callable[[], None]) -> None:
        if not self._closing:
            self._ui_queue.put(callback)

    def _drain_ui_queue(self) -> None:
        if self._closing:
            return
        for _ in range(100):
            try:
                callback = self._ui_queue.get_nowait()
            except queue.Empty:
                break
            try:
                callback()
            except Exception:
                pass
        self.after(100 if self.performance.low_memory else 60, self._drain_ui_queue)

    def _configure_styles(self):
        c = self.colors
        style = ttk.Style(self)
        try:
            style.theme_use("clam")
        except tk.TclError:
            pass

        style.configure(".", background=c["panel"], foreground=c["text"], fieldbackground=c["field"], bordercolor=c["border"], lightcolor=c["border"], darkcolor=c["border"])
        style.configure("TFrame", background=c["panel"])
        style.configure("Page.TFrame", background=c["bg"])
        style.configure("Card.TFrame", background=c["panel"], borderwidth=1, relief="solid")
        style.configure("Field.TFrame", background=c["field"], borderwidth=1, relief="solid")
        style.configure("Sidebar.TFrame", background=c["sidebar"])

        style.configure("TLabel", background=c["panel"], foreground=c["text"])
        style.configure("PageTitle.TLabel", font=("Segoe UI Semibold", 24), background=c["bg"], foreground=c["text"])
        style.configure("LoginTitle.TLabel", font=("Segoe UI Semibold", 30), background=c["bg"], foreground=c["text"])
        style.configure("SectionTitle.TLabel", font=("Segoe UI Semibold", 14), background=c["bg"], foreground=c["text"])
        style.configure("DialogTitle.TLabel", font=("Segoe UI Semibold", 14), background=c["panel"], foreground=c["text"])
        style.configure("Muted.TLabel", foreground=c["muted"], background=c["bg"])
        style.configure("CardMuted.TLabel", foreground=c["muted"], background=c["panel"])
        style.configure("CardSectionTitle.TLabel", font=("Segoe UI Semibold", 14), background=c["panel"], foreground=c["text"])
        style.configure("RoleBadge.TLabel", font=("Segoe UI Semibold", 9), background=c["accent_soft"], foreground=c["accent_strong"], padding=(10, 5))
        style.configure("ErrorCard.TLabel", foreground=c["danger"], background=c["panel"])
        style.configure("Metric.TLabel", font=("Segoe UI Semibold", 18), background=c["panel"], foreground=c["accent_strong"])
        style.configure("Timer.TLabel", font=("Consolas", 16, "bold"), foreground=c["accent_strong"], background=c["bg"])
        style.configure("CardTimer.TLabel", font=("Consolas", 16, "bold"), foreground=c["accent_strong"], background=c["panel"])
        style.configure("Code.TLabel", font=("Consolas", 28, "bold"), foreground=c["accent_strong"], background=c["panel"])

        style.configure("TEntry", fieldbackground=c["field"], foreground=c["text"], insertcolor=c["text"], borderwidth=1, padding=9)
        style.configure("TCombobox", fieldbackground=c["field"], foreground=c["text"], arrowcolor=c["text"], padding=8)
        style.map("TCombobox", fieldbackground=[("readonly", c["field"])], foreground=[("readonly", c["text"])])
        style.configure("TSpinbox", fieldbackground=c["field"], foreground=c["text"], arrowcolor=c["text"], padding=8)
        # Native ttk buttons are retained as a fallback; visible application buttons use RoundedButton.
        style.configure("TButton", padding=(13, 8), borderwidth=0, font=("Segoe UI Semibold", 10))

        style.configure("Treeview", background=c["panel"], fieldbackground=c["panel"], foreground=c["text"], rowheight=32, bordercolor=c["border"])
        style.map("Treeview", background=[("selected", c["accent_soft"])], foreground=[("selected", c["text"])])
        style.configure("Treeview.Heading", background=c["panel2"], foreground=c["text"], padding=9, font=("Segoe UI Semibold", 9))
        style.configure("TNotebook", background=c["bg"], borderwidth=0)
        style.configure("TNotebook.Tab", background=c["panel2"], foreground=c["muted"], padding=(16, 9))
        style.map("TNotebook.Tab", background=[("selected", c["accent_soft"])], foreground=[("selected", c["accent_strong"])])
        style.configure("TLabelframe", background=c["panel"], foreground=c["text"], bordercolor=c["border"])
        style.configure("TLabelframe.Label", background=c["panel"], foreground=c["text"], font=("Segoe UI Semibold", 10))
        style.configure("TCheckbutton", background=c["panel"], foreground=c["text"])

    def _build_root_views(self):
        self.columnconfigure(0, weight=1)
        self.rowconfigure(0, weight=1)

        self.login_page = LoginPage(self, self)
        self.login_page.grid(row=0, column=0, sticky="nsew")

        self.shell = ttk.Frame(self, style="Page.TFrame")
        self.shell.grid(row=0, column=0, sticky="nsew")
        self.shell.columnconfigure(1, weight=1)
        self.shell.rowconfigure(0, weight=1)

        self.sidebar = ttk.Frame(self.shell, style="Sidebar.TFrame", padding=(12, 18))
        self.sidebar.grid(row=0, column=0, sticky="ns")
        self.sidebar.configure(width=228)
        self.sidebar.grid_propagate(False)

        self.content = ttk.Frame(self.shell, style="Page.TFrame")
        self.content.grid(row=0, column=1, sticky="nsew")
        self.content.columnconfigure(0, weight=1)
        self.content.rowconfigure(0, weight=1)

    def _clear_role_ui(self):
        for child in self.sidebar.winfo_children():
            child.destroy()
        for child in self.content.winfo_children():
            child.destroy()
        self.pages = {}
        self.nav_buttons = {}

    def _build_role_ui(self):
        self._clear_role_ui()
        c = self.colors
        ttk.Label(self.sidebar, text="DT", font=("Segoe UI Semibold", 25), foreground=c["accent_strong"], background=c["sidebar"]).pack(anchor="w", padx=10)
        ttk.Label(self.sidebar, text="DictaType", font=("Segoe UI Semibold", 12), foreground=c["text"], background=c["sidebar"]).pack(anchor="w", padx=10)

        if self.current_role == "teacher":
            ttk.Label(self.sidebar, text="Teacher workspace", foreground=c["muted"], background=c["sidebar"]).pack(anchor="w", padx=10, pady=(3, 22))
            nav = [
                ("teacher", "Dashboard"),
                ("classroom", "Local classroom"),
                ("settings", "Settings & security"),
            ]
            self.pages = {
                "teacher": TeacherPage(self, self.content),
                "classroom": ClassroomPage(self, self.content),
                "settings": SettingsPage(self, self.content),
            }
            initial = "teacher"
        else:
            name = (self.current_student or {}).get("name", "Student")
            ttk.Label(self.sidebar, text=name, foreground=c["text"], background=c["sidebar"], font=("Segoe UI Semibold", 10)).pack(anchor="w", padx=10, pady=(3, 0))
            class_name = (self.current_student or {}).get("class_name", "")
            if class_name:
                ttk.Label(self.sidebar, text=class_name, foreground=c["muted"], background=c["sidebar"]).pack(anchor="w", padx=10, pady=(1, 20))
            else:
                ttk.Label(self.sidebar, text="Student workspace", foreground=c["muted"], background=c["sidebar"]).pack(anchor="w", padx=10, pady=(1, 20))
            nav = [("student", "Practice")]
            self.pages = {"student": StudentPage(self, self.content)}
            initial = "student"

        for page in self.pages.values():
            page.grid(row=0, column=0, sticky="nsew")

        for key, label in nav:
            button = RoundedButton(self.sidebar, text=label, style="Sidebar.TButton", width=24, command=lambda k=key: self.show_page(k))
            button.pack(fill="x", pady=3)
            self.nav_buttons[key] = button

        bottom = ttk.Frame(self.sidebar, style="Sidebar.TFrame")
        bottom.pack(side="bottom", fill="x", pady=(10, 0))
        RoundedButton(bottom, text="Sign out", style="Secondary.TButton", command=self.logout).pack(fill="x")
        ttk.Label(bottom, text=f"Open source · v{APP_VERSION}", foreground=c["muted"], background=c["sidebar"]).pack(anchor="w", padx=10, pady=(12, 0))
        self.show_page(initial)

    def show_login(self):
        self.current_role = None
        self.current_student = None
        self.teacher_authenticated = False
        self.login_page.reset()
        self.login_page.tkraise()

    def _lock_status(self, key: str) -> int:
        remaining = int(self._auth_locked_until.get(key, 0) - time.monotonic())
        if remaining <= 0:
            self._auth_locked_until.pop(key, None)
            return 0
        return remaining

    def _auth_failed(self, key: str) -> str:
        count = self._auth_failures.get(key, 0) + 1
        self._auth_failures[key] = count
        if count >= 5:
            self._auth_failures[key] = 0
            self._auth_locked_until[key] = time.monotonic() + 60
            return "Too many incorrect attempts. This login is locked for 60 seconds."
        remaining = 5 - count
        return f"Incorrect login details. {remaining} attempt(s) remain before a temporary lock."

    def _auth_success(self, key: str):
        self._auth_failures.pop(key, None)
        self._auth_locked_until.pop(key, None)

    def login_teacher(self, pin: str) -> tuple[bool, str]:
        key = "teacher"
        locked = self._lock_status(key)
        if locked:
            return False, f"Teacher login is temporarily locked. Try again in {locked} seconds."
        if not self.db.check_pin(pin):
            return False, self._auth_failed(key)
        self._auth_success(key)
        if self.db.get_setting("security_setup_complete", "0") != "1":
            if not self._complete_teacher_security_setup():
                return False, "Teacher security setup must be completed before access is granted."
        self.start_session("teacher")
        return True, ""

    def _complete_teacher_security_setup(self) -> bool:
        messagebox.showinfo(
            "Secure teacher account",
            "Before using teacher features, replace the initial PIN with your own 6 to 12 digit PIN.",
            parent=self,
        )
        new = simpledialog.askstring("New teacher PIN", "Choose a 6 to 12 digit teacher PIN:", show="●", parent=self)
        if new is None:
            return False
        if not new.isdigit() or not 6 <= len(new) <= 12:
            messagebox.showerror("Invalid PIN", "The teacher PIN must contain 6 to 12 digits.", parent=self)
            return False
        confirm = simpledialog.askstring("Confirm teacher PIN", "Enter the new PIN again:", show="●", parent=self)
        if confirm != new:
            messagebox.showerror("PIN mismatch", "The PINs do not match.", parent=self)
            return False
        self.db.set_pin(new)
        self.db.set_setting("security_setup_complete", "1")
        return True

    def login_student(self, identifier: str) -> tuple[bool, str]:
        identifier = identifier.strip()
        if not identifier:
            return False, "Choose a student profile or create a new one."
        student = self.db.authenticate_student(identifier)
        if not student:
            return False, "That student profile is no longer available. Ask the teacher to re-enable it or create a new profile."
        self.start_session("student", student)
        return True, ""

    def login_student_by_name(self, name: str, class_name: str = "") -> tuple[bool, str]:
        name = name.strip()
        class_name = class_name.strip()
        if not name:
            return False, "Enter your name."
        student = self.db.find_student(name, class_name)
        if not student:
            return False, "No matching profile was found. Check your name and class, or choose Create my profile."
        if not bool(student.get("active", 1)):
            return False, "This student profile has been disabled by the teacher."
        self.start_session("student", student)
        return True, ""

    def start_session(self, role: str, student: dict[str, Any] | None = None):
        self.current_role = role
        self.current_student = student
        self.teacher_authenticated = role == "teacher"
        self.touch_activity()
        self._build_role_ui()
        self.shell.tkraise()
        self.refresh_all()

    def logout(self, silent: bool = False):
        self.attributes("-fullscreen", False)
        self.speech.stop()
        if self.current_role == "teacher":
            self.classroom_server.stop()
        self._clear_role_ui()
        self.show_login()
        if not silent:
            self.login_page.error_label.configure(text="Signed out securely.")

    def show_page(self, key: str):
        if key not in self.pages:
            return
        page = self.pages[key]
        page.tkraise()
        for name, button in self.nav_buttons.items():
            button.state(["selected"] if name == key else ["!selected"])
        refresh = getattr(page, "refresh", None)
        if callable(refresh):
            refresh()
        self.touch_activity()

    def refresh_all(self):
        for page in self.pages.values():
            refresh = getattr(page, "refresh", None)
            if callable(refresh):
                try:
                    refresh()
                except Exception:
                    pass

    def touch_activity(self, _event=None):
        self.last_activity = time.monotonic()

    def _security_tick(self):
        if self._closing:
            return
        if self.current_role == "teacher":
            try:
                minutes = int(self.db.get_setting("teacher_auto_lock_minutes", "15"))
            except ValueError:
                minutes = 15
            if minutes > 0 and time.monotonic() - self.last_activity >= minutes * 60:
                self.logout(silent=True)
                messagebox.showinfo("Teacher session locked", "The teacher session was locked after inactivity.", parent=self)
        self.after(30_000, self._security_tick)

    def voices_for_language(self, language: str) -> list[Voice]:
        language = language.casefold()
        matches = []
        for voice in self.voices:
            haystack = " ".join([voice.id, voice.name, *voice.languages]).casefold()
            if language == "fr" and any(token in haystack for token in ["fr-", "fr_", "french", "français", "hortense", "denise"]):
                matches.append(voice)
            elif language == "en" and any(token in haystack for token in ["en-", "en_", "english", "zira", "david", "mark", "hazel"]):
                matches.append(voice)
        if matches:
            return matches
        # Never offer the dedicated French neural model for an English lesson.
        return [voice for voice in self.voices if not voice.id.startswith("dictatype:piper:")]

    def refresh_voices(self):
        self.voices = list_voices()
        self.refresh_all()

    def on_close(self):
        self._closing = True
        try:
            self.db.set_setting("window_geometry", self.geometry())
        except Exception:
            pass
        self.speech.stop()
        self.classroom_server.stop()
        self.destroy()

